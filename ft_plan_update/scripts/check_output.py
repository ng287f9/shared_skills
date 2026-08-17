#!/usr/bin/env python3
"""Check FT_Plan_Update output file for issues."""
import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

path = sys.argv[1]
print("Checking: %s" % path)
wb = openpyxl.load_workbook(path)
ws = wb.active

# 1. PCT
print("\n=== PCT (Col 2) first 10 ===")
for r in range(2, 12):
    v = ws.cell(r, 2).value
    t = type(v).__name__
    print("Row %d: val=%s type=%s" % (r, v, t))

# 2. Red font in Col 3
print("\n=== Red font cells ===")
red = 0
for r in range(2, ws.max_row+1):
    cell = ws.cell(r, 3)
    fc = cell.font.color
    if fc and fc.rgb:
        try:
            rgb = str(fc.rgb)
            ri = int(rgb[2:4], 16)
            gi = int(rgb[4:6], 16)
            bi = int(rgb[6:8], 16)
            if ri > 200 and gi < 80 and bi < 80:
                red += 1
                print("Row %d: fg=%s [%s]" % (r, rgb, str(cell.value)[:50]))
        except:
            pass
print("Total red in Col 3: %d" % red)

# 3. Untranslated columns
print()
for label, col in [("Prerequisite",7),("RMKS/Notes",8),("Paperwork",9),("Reason",14),("AddReq",15)]:
    eng = 0
    for r in range(2, ws.max_row+1):
        v = str(ws.cell(r, col).value or "")
        if v.strip() and any(c.isalpha() and ord(c) < 128 for c in v):
            eng += 1
    print("%s (Col%d): %d English" % (label, col, eng))

wb.close()
