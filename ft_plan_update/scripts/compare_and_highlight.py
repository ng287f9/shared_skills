#!/usr/bin/env python3
"""
Phase 4 (Optional): Compare new Excel with previous version and highlight changes.

Usage:
    python scripts/compare_and_highlight.py \
        --new "FT_Plan_Update_20260708_220449.xlsx" \
        --previous "previous.xlsx" \
        --output "FT_Plan_Update_20260708_220449.xlsx"

Compares row-by-row and cell-by-cell:
- New rows: highlight entire row in light gray
- Deleted rows: (skipped, as they're not in new file)
- Modified cells: highlight cell in light gray
- Saves updated workbook with change highlighting
"""

import argparse
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("Error: openpyxl is required", file=sys.stderr)
    sys.exit(1)


# Gray fill for changed cells
CHANGE_FILL = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")


def normalize_value(val):
    """Normalize cell value for comparison (handles None, spaces, etc.)"""
    if val is None:
        return ""
    val_str = str(val).strip()
    return val_str


def compare_workbooks(new_path, prev_path):
    """
    Compare two workbooks and identify changes.
    Returns dict with change info: {row: {col: (old_val, new_val), ...}, ...}
    """
    # Load workbooks
    try:
        wb_new = openpyxl.load_workbook(new_path)
        ws_new = wb_new.active
    except Exception as e:
        print(f"Error loading new workbook: {e}", file=sys.stderr)
        return None

    try:
        wb_prev = openpyxl.load_workbook(prev_path, data_only=True)
        ws_prev = wb_prev.active
    except Exception as e:
        print(f"Warning: Could not load previous version: {e}", file=sys.stderr)
        return None

    changes = {}  # {row: {col: (old, new), ...}}

    # Determine max rows/columns
    max_row_new = ws_new.max_row
    max_row_prev = ws_prev.max_row
    max_col = max(ws_new.max_column, ws_prev.max_column)

    # Compare rows
    max_row = max(max_row_new, max_row_prev)

    for row in range(1, max_row + 1):
        row_changes = {}

        for col in range(1, max_col + 1):
            # Get values
            new_val = None
            old_val = None

            if row <= max_row_new:
                new_cell = ws_new.cell(row, col)
                new_val = normalize_value(new_cell.value)

            if row <= max_row_prev:
                old_cell = ws_prev.cell(row, col)
                old_val = normalize_value(old_cell.value)

            # Check if changed
            if new_val != old_val:
                row_changes[col] = (old_val, new_val)

        if row_changes:
            changes[row] = row_changes

    wb_prev.close()
    return changes, ws_new, wb_new


def apply_highlighting(ws, workbook, changes):
    """Apply gray highlighting to changed cells in the worksheet."""
    highlighted_count = 0

    for row, col_changes in changes.items():
        for col, (old_val, new_val) in col_changes.items():
            cell = ws.cell(row, col)

            # Preserve existing formatting, only add background color
            existing_fill = cell.fill
            if existing_fill and existing_fill.fill_type != 'solid':
                # Keep existing fill if not solid
                pass
            else:
                # Apply gray background
                cell.fill = CHANGE_FILL
                highlighted_count += 1

    return highlighted_count


def main():
    parser = argparse.ArgumentParser(description="Compare and highlight changes between Excel versions")
    parser.add_argument("--new", required=True, help="Path to new .xlsx file")
    parser.add_argument("--previous", required=True, help="Path to previous .xlsx file")
    parser.add_argument("--output", required=True, help="Output .xlsx path (can be same as --new)")
    args = parser.parse_args()

    if not os.path.exists(args.new):
        print(f"Error: New file not found: {args.new}", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(args.previous):
        print(f"Warning: Previous file not found: {args.previous}")
        print("Skipping comparison (file will be saved unchanged)")
        # Just copy new file to output
        import shutil
        shutil.copy2(args.new, args.output)
        print(f"Output saved to: {args.output}")
        return

    print(f"Comparing:")
    print(f"  New:      {args.new}")
    print(f"  Previous: {args.previous}")
    print()

    # Compare workbooks
    result = compare_workbooks(args.new, args.previous)
    if result is None:
        print("Comparison failed. Output saved unchanged.")
        import shutil
        shutil.copy2(args.new, args.output)
        sys.exit(1)

    changes, ws_new, wb_new = result

    print(f"Changes detected: {len(changes)} rows")
    total_cells_changed = sum(len(cols) for cols in changes.values())
    print(f"  Total cells changed: {total_cells_changed}")

    # Show samples of changes
    sample_count = 0
    for row in sorted(changes.keys())[:5]:
        col_changes = changes[row]
        if col_changes:
            for col, (old_val, new_val) in list(col_changes.items())[:2]:
                old_display = old_val[:30] if old_val else "(empty)"
                new_display = new_val[:30] if new_val else "(empty)"
                print(f"    Row {row}, Col {col}: '{old_display}' → '{new_display}'")
                sample_count += 1
                if sample_count >= 3:
                    break
        if sample_count >= 3:
            break

    # Apply highlighting
    print()
    print("Applying gray highlighting to changed cells...")
    highlighted = apply_highlighting(ws_new, wb_new, changes)
    print(f"  {highlighted} cells highlighted")

    # Save
    print()
    print(f"Saving: {args.output}")
    wb_new.save(args.output)
    wb_new.close()
    print("Done!")


if __name__ == "__main__":
    main()
