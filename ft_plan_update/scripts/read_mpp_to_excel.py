#!/usr/bin/env python3
"""
Phase 2: Build Excel directly from MPP data with MPP formatting.

Usage:
    python scripts/read_mpp_to_excel.py \
        --mpp FlightTestSchedule_20260626_PlanB.mpp \
        --template 术语对照_中文.xlsx \
        --output "FT_Plan_Update_20260708.xlsx"

Strategy:
    - Reads ALL data from the MPP file via mpxj (19-column Entry table)
    - Extracts per-cell formatting from GanttChartView.ColumnFontStyle
    - Extracts column header names from 术语对照_中文.xlsx (row 1 only)
    - Builds a new Excel matching MPP appearance
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required", file=sys.stderr)
    sys.exit(1)


# ── Default MPP Gantt Chart View style constants ──
# (from GanttChartView string output)
FONT_NAME = "Calibri"
FONT_SIZE_NORMAL = 11
FONT_SIZE_PCT = 10
FONT_SIZE_HEADER = 10
FONT_SIZE_PROJECT_SUMMARY = 12

# Grid line colour from MPP: SheetRowsGridLines (218,220,221)
GRID_COLOR = "DADCDD"

# Default fills
HEADER_FILL = PatternFill(start_color="FFDADCDD", end_color="FFDADCDD", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

# Border style matching MPP grid lines
THIN_SIDE = Side(style="thin", color=GRID_COLOR)
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE,
)

# ── MPP Table Column → Excel Column mapping ──
# Based on the Entry table parsed via MPXJ
COL_MAP = [
    # (excel_col, header_key, mpp_field, format_func)
    ("ID", "getID", None),
    ("% Complete", "getPercentageComplete", "pct"),
    ("Task Name", "getName", "name"),
    ("Planned A/C", "getText(1)", "text1"),
    ("Required Sortie#", "getNumber(1)", "number1"),
    ("Required FH", "getNumber(2)", "number2"),
    ("Prerequisite", "getText(16)", "text16"),
    ("RMKS", "getText(12)", "text12"),
    ("Paperwork status", "getText(26)", "text26"),
    ("Duration", "getDuration", "duration"),
    ("Start", "getStart", "start"),
    ("Finish", "getFinish", "finish"),
    ("Predecessors", "getPredecessors", "predecessors"),
    ("Reason for specific A/C", "getText(10)", "text10"),
    ("Additional Requirements", "getText(18)", "text18"),
    ("Actual Start", "getActualStart", "actual_start"),
    ("Actual Finish", "getActualFinish", "actual_finish"),
]


def rgb_to_hex(r, g, b):
    """Convert RGB color tuple to hex string for openpyxl."""
    return "FF%02X%02X%02X" % (int(r) & 0xFF, int(g) & 0xFF, int(b) & 0xFF)


def try_parse_mpp(mpp_path):
    """
    Read MPP file via mpxj/JPype.
    Returns: (tasks_list, col_styles_dict)
      tasks_list: list of task dicts with all fields
      col_styles: dict (row_index, field_name) -> style_info
    """
    try:
        import mpxj
        import jpype

        if not jpype.isJVMStarted():
            if "JAVA_HOME" not in os.environ:
                for jhome in [
                    r"C:\Program Files\Eclipse Adoptium\jdk-17.0.17.10-hotspot",
                    r"C:\Program Files\Eclipse Adoptium\jdk-21.0.9.10-hotspot",
                    r"C:\Program Files\Java\jdk-17",
                    r"C:\Program Files\Java\jdk-21",
                ]:
                    if os.path.isdir(jhome):
                        os.environ["JAVA_HOME"] = jhome
                        break

            jvmpath = jpype.getDefaultJVMPath()
            jar_dir = mpxj.dirpath
            jars = [os.path.join(jar_dir, f) for f in sorted(os.listdir(jar_dir)) if f.endswith(".jar")]
            jpype.startJVM(jvmpath, classpath=";".join(jars))

        from org.mpxj.reader import UniversalProjectReader
        proj = UniversalProjectReader().read(mpp_path)

        # ── 1. Extract task data ──
        uid_to_idx = {}  # uniqueID -> 0-based row index
        tasks_list = []
        idx = 0

        for t in proj.getTasks():
            ol = t.getOutlineLevel()
            if ol is None or ol == 0:
                continue

            uid = int(t.getUniqueID())
            uid_to_idx[uid] = idx

            # ── Duration formatting ──
            dur_obj = t.getDuration()
            if dur_obj:
                dur_val = dur_obj.getDuration()
                dur_units = str(dur_obj.getUnits())
                unit_labels = {"d": " 天", "w": " 周", "h": " 小时", "mo": " 月"}
                dur_text = "%.4g%s" % (dur_val, unit_labels.get(dur_units, " " + dur_units))
            else:
                dur_text = ""

            # ── Date formatting ──
            def fmt_date(dt):
                if dt is None:
                    return ""
                dt_str = str(dt)
                try:
                    d = datetime.strptime(dt_str[:10], "%Y-%m-%d")
                    return "%d年%d月%d日" % (d.year, d.month, d.day)
                except (ValueError, IndexError):
                    return dt_str[:10]

            start_text = fmt_date(t.getStart())
            finish_text = fmt_date(t.getFinish())
            actual_start_text = fmt_date(t.getActualStart())
            actual_finish_text = fmt_date(t.getActualFinish())

            # ── % Complete (raw 0-100 → 0-1 decimal) ──
            pct_obj = t.getPercentageComplete()
            # MPP stores %Complete as 0-100; store as 0-1 decimal with percentage format
            pct_val = float(pct_obj) / 100.0 if pct_obj is not None else 0.0

            # ── Predecessors ──
            pred_ids = []
            preds = t.getPredecessors()
            if preds:
                for p in preds:
                    src = p.getPredecessorTask()
                    if src:
                        pred_ids.append(str(src.getUniqueID()))

            # ── Custom field helpers ──
            def safe_text(field_num):
                try:
                    v = t.getText(field_num)
                    return str(v).strip() if v else ""
                except:
                    return ""

            def safe_number(field_num):
                try:
                    return t.getNumber(field_num)
                except:
                    return None

            task = {
                "id": int(t.getID()),
                "uid": uid,
                "name": str(t.getName() or "").strip(),
                "pct": pct_val,
                "duration": dur_text,
                "start": start_text,
                "finish": finish_text,
                "predecessors": ", ".join(pred_ids),
                "text1": safe_text(1),       # Planned A/C
                "text10": safe_text(10),     # Reason for specific A/C
                "text12": safe_text(12),     # RMKS
                "text16": safe_text(16),     # Prerequisite
                "text18": safe_text(18),     # Additional Requirements
                "text26": safe_text(26),     # Paperwork status
                "number1": safe_number(1),   # Required Sortie#
                "number2": safe_number(2),   # Required FH
                "actual_start": actual_start_text,
                "actual_finish": actual_finish_text,
                "notes": str(t.getNotes() or "").strip(),
                "outline_level": ol,
                "summary": bool(t.getSummary()),
                "milestone": bool(t.getMilestone()),
            }
            tasks_list.append(task)
            idx += 1

        # ── 2. Extract ColumnFontStyle data ──
        col_styles = {}  # (row_idx, field_name) -> {bg, fg, bold, font_size}
        red_rows = set()  # row indices that have ANY red-bg ColumnFontStyle entry

        views = proj.getViews()
        for v in views:
            s = str(v)
            if "GanttChartView" not in s:
                continue

            lines = s.split("\n")
            for line in lines:
                line = line.strip()
                if "ColumnFontStyle" not in line:
                    continue

                # Parse rowUniqueID
                m = re.search(r'rowUniqueID=(\d+)', line)
                if not m:
                    continue
                style_uid = int(m.group(1))
                if style_uid not in uid_to_idx:
                    # Could be resource UID or other — skip
                    continue
                row_idx = uid_to_idx[style_uid]

                # Parse fieldType
                m = re.search(r'fieldType=(\S+)', line)
                ft = m.group(1).rstrip("]") if m else None

                # Determine which column(s) to apply to
                # fieldType "null" means apply to all cells in this row
                # fieldType like "Text26", "Text12", "ID", "% Complete" etc

                # Parse font
                m = re.search(r'FontBase name=(\S+) size=(\d+)', line)
                font_size = int(m.group(2)) if m else None

                bold = "bold=true" in line

                # Parse text color
                m = re.search(r'color=java\.awt\.Color\[r=(\d+),g=(\d+),b=(\d+)\]', line)
                fg = (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

                # Parse background color
                m = re.search(r'backgroundColor=java\.awt\.Color\[r=(\d+),g=(\d+),b=(\d+)\]', line)
                bg = (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

                style_info = {"row_idx": row_idx, "ft": ft,
                              "bold": bold, "bg": bg, "fg": fg,
                              "font_size": font_size}

                # When bg is red (255,0,0), convert to red font instead
                # This marks "新增项目/新加内容" items in the output
                if bg and bg[0] >= 250 and bg[1] < 50 and bg[2] < 50:
                    style_info["fg"] = bg      # red font
                    style_info["bg"] = None    # no red background
                    red_rows.add(row_idx)

                # Map fieldType to our column names
                ft_mapping = {
                    "null": None,           # all columns
                    "ID": "id",
                    "Task": "name",         # Task Name field (short)
                    "Task Name": "name",     # Task Name field
                    "%": "pct",             # % Complete (short)
                    "% Complete": "pct",
                    "Text1": "text1",
                    "Text10": "text10",
                    "Text12": "text12",
                    "Text16": "text16",
                    "Text18": "text18",
                    "Text26": "text26",
                    "Start": "start",
                    "Finish": "finish",
                    "Duration": "duration",
                    "Number1": "number1",
                    "Number2": "number2",
                }

                col_key = ft_mapping.get(ft)
                if col_key:
                    if col_key not in col_styles:
                        col_styles[col_key] = {}
                    col_styles[col_key][row_idx] = style_info
                elif ft is None or ft == "null":
                    # Apply to all columns
                    for ck in ["id", "pct", "name", "text1", "number1", "number2",
                               "text16", "text12", "text26", "duration", "start", "finish",
                               "predecessors", "text10", "text18", "actual_start", "actual_finish"]:
                        if ck not in col_styles:
                            col_styles[ck] = {}
                        if row_idx not in col_styles[ck]:
                            col_styles[ck][row_idx] = style_info

        return tasks_list, col_styles, red_rows

    except Exception as e:
        print("Error reading MPP: %s" % e, file=sys.stderr)
        import traceback
        traceback.print_exc()
        return None, None, set()


def read_headers_from_template(template_path):
    """Read column headers from row 1 of the Chinese template."""
    if not os.path.exists(template_path):
        print("Warning: template %s not found, using default headers" % template_path)
        return [
            "任务号", "完成百分比", "任务名称", "计划飞机",
            "需要的架次", "需要的飞行时间", "前提条件", "备注",
            "文件状态", "工期", "开始时间", "完成时间", "前置任务",
            "使用特定飞机原因", "额外需求", "实际开始时间", "实际完成时间",
        ]

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        headers.append(str(v).strip() if v else "")
    wb.close()
    print("  Read %d headers from template" % len(headers))
    return headers


def build_excel(tasks_list, col_styles, red_rows, headers, output_path):
    """Build Excel from MPP data with MPP formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flight Test Plan"

    num_rows = len(tasks_list)
    num_cols = len(headers)

    # ── Write headers (Row 1) ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE_HEADER, bold=False, color="FF363636")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # ── Set column widths ──
    widths = [6, 10, 50, 12, 10, 14, 20, 24, 13, 12, 16, 16, 10, 22, 20, 16, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Freeze top row ──
    ws.freeze_panes = "A2"

    # ── Field → column index mapping ──
    field_to_col = {
        "id": 1, "pct": 2, "name": 3, "text1": 4,
        "number1": 5, "number2": 6, "text16": 7, "text12": 8,
        "text26": 9, "duration": 10, "start": 11, "finish": 12,
        "predecessors": 13, "text10": 14, "text18": 15,
        "actual_start": 16, "actual_finish": 17,
    }

    # ── Write data rows ──
    for row_idx, task in enumerate(tasks_list):
        excel_row = row_idx + 2

        # ── Task name with indentation ──
        level = task["outline_level"]
        indent = "  " * (level - 1) if level > 1 else ""
        display_name = indent + task["name"]

        # ── Cell values per column ──
        values = {
            "id": task["id"],
            "pct": round(task["pct"], 2),
            "name": display_name,
            "text1": task["text1"] or "",
            "number1": task["number1"],
            "number2": task["number2"],
            "text16": task["text16"] or "",
            "text12": task["text12"] or task["notes"] or "",
            "text26": task["text26"] or "",
            "duration": task["duration"],
            "start": task["start"],
            "finish": task["finish"],
            "predecessors": task["predecessors"],
            "text10": task["text10"] or "",
            "text18": task["text18"] or "",
            "actual_start": task["actual_start"],
            "actual_finish": task["actual_finish"],
        }

        for field, col in field_to_col.items():
            cell = ws.cell(excel_row, col, values[field])

            # ── Determine style for this cell ──
            bg = None
            fg = None
            is_bold = task["summary"]  # Default: bold for summary rows
            font_size = FONT_SIZE_NORMAL

            # ColumnFontStyle overrides
            if col_styles and field in col_styles:
                field_styles = col_styles[field]
                if row_idx in field_styles:
                    style = field_styles[row_idx]
                    if style.get("bg"):
                        bg = style["bg"]
                    if style.get("fg"):
                        fg = style["fg"]
                    if "bold" in style:
                        is_bold = style["bold"]
                    if style.get("font_size"):
                        font_size = style["font_size"]

            # Mark new items (red_rows) with red font
            if red_rows and row_idx in red_rows:
                fg = (255, 0, 0)  # Red color for new items

            # % Complete column always uses 10pt (MPP default) and percentage format
            if field == "pct":
                font_size = FONT_SIZE_PCT
                cell.number_format = "0%"

            # ── Build font ──
            font_color = "FF000000"
            if fg:
                font_color = rgb_to_hex(*fg)
            font = Font(name=FONT_NAME, size=font_size, bold=is_bold, color=font_color)

            # ── Build fill ──
            if bg:
                fill = PatternFill(start_color=rgb_to_hex(*bg), end_color=rgb_to_hex(*bg), fill_type="solid")
            else:
                fill = WHITE_FILL

            cell.font = font
            cell.fill = fill
            cell.border = THIN_BORDER

            # ── Alignment ──
            if col == 3:
                # Task name: left, wrap
                halign = "left"
                vwrap = True
            elif field in ("id", "pct"):
                halign = "center"
                vwrap = False
            elif field in ("number1", "number2"):
                halign = "right"
                vwrap = False
            elif field in ("start", "finish", "actual_start", "actual_finish", "duration"):
                halign = "center"
                vwrap = False
            else:
                halign = "left"
                vwrap = True

            cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=vwrap)

        # Adjust row height for wrapped content
        if task.get("text12") or task.get("notes"):
            ws.row_dimensions[excel_row].height = 30

    # ── Auto-filter ──
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = "A1:%s%d" % (last_col_letter, num_rows + 1)

    wb.save(output_path)
    wb.close()
    print("  Written %d rows to %s" % (num_rows, output_path))
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Build Excel from MPP data with MPP formatting")
    parser.add_argument("--mpp", required=True, help="Path to .mpp file")
    parser.add_argument("--template", default="术语对照_中文.xlsx", help="Chinese template (for headers only)")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    if not os.path.exists(args.mpp):
        print("Error: MPP file not found: %s" % args.mpp, file=sys.stderr)
        sys.exit(1)

    print("Reading MPP file: %s" % args.mpp)
    tasks_list, col_styles, red_rows = try_parse_mpp(args.mpp)

    if tasks_list is None:
        print("Error: Could not read MPP file. Java 17+ is required.", file=sys.stderr)
        sys.exit(1)

    print("  Read %d tasks, %d column style groups" % (len(tasks_list), len(col_styles)))

    headers = read_headers_from_template(args.template)

    build_excel(tasks_list, col_styles, red_rows, headers, args.output)

    print("Done!")


if __name__ == "__main__":
    main()
