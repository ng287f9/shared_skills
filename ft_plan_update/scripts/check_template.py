"""Quick check of Chinese template"""
import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active

# Check if any data row has "新加内容" or similar text in Col 3
for r in range(2, ws.max_row + 1):
    v = str(ws.cell(r, 3).value or "")
    if "新加" in v or "新增" in v:
        print("Row%d: [%s]" % (r, v[:60]))

# Check all rows with red font
print()
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, 3)
    fc = cell.font.color
    if fc and fc.rgb and len(str(fc.rgb)) >= 8:
        try:
            ri = int(str(fc.rgb)[2:4], 16)
            gi = int(str(fc.rgb)[4:6], 16)
            bi = int(str(fc.rgb)[6:8], 16)
            if ri > 200 and gi < 80 and bi < 80:
                print("Row%d: RED font [%s]" % (r, str(cell.value)[:50]))
        except:
            pass

print()
# Check Col 8, 9, 14, 15 for any red font
for col, label in [(8,"RMKS"),(9,"Paperwork"),(14,"Reason"),(15,"AddReq")]:
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, col)
        fc = cell.font.color
        if fc and fc.rgb and len(str(fc.rgb)) >= 8:
            try:
                ri = int(str(fc.rgb)[2:4], 16)
                if ri > 200:
                    print("Row%d Col%d: RED font [%s]" % (r, col, str(cell.value)[:40]))
            except:
                pass

wb.close()
