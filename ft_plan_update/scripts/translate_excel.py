#!/usr/bin/env python3
"""
Phase 3: Translate English text in the Excel to Chinese.

Usage:
    python scripts/translate_excel.py \
        --input "FT_Plan_Update_20260708.xlsx" \
        --glossary references/术语对照.md \
        --output "FT_Plan_Update_20260708.xlsx"

Strategy:
    - Column 3 (Task Name): glossary lookup (exact → fuzzy → FRQ-stripped → keyword → builtin dict)
    - Columns 7,8,9,14,15 (Prerequisite, RMKS, Paperwork, Reason, Additional Req): builtin dict
    - If found: replaces cell value with Chinese text
    - If not found: keeps original English
    - Preserves existing cell formatting
"""

import argparse
import difflib
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("Error: openpyxl is required", file=sys.stderr)
    sys.exit(1)


# ── Built-in aviation/aerospace phrase dictionary ──
BUILTIN_DICT = {
    # Task names
    "flight test": "试飞",
    "flight test campaign": "试飞活动",
    "certification": "取证",
    "certification flight test": "取证试飞",
    "inspection": "检查",
    "maintenance": "维护",
    "installation": "安装",
    "ground testing": "地面测试",
    "ground test": "地面测试",
    "functional test": "功能测试",
    "functional testing": "功能测试",
    "performance check": "性能检查",
    "calibration": "校准",
    "verification": "验证",
    "evaluation": "评估",
    "test": "测试",
    "system": "系统",
    "software": "软件",
    "delivery": "交付",
    "development": "研发",
    "removal": "拆除",
    "ferry flight": "调机飞行",
    "water ops": "水上运行",
    "stall": "失速",
    "stall warning": "失速警告",
    "take-off": "起飞",
    "takeoff": "起飞",
    "landing": "着陆",
    "climb": "爬升",
    "descent": "下降",
    "trim": "配平",
    "fuel": "燃油",
    "engine": "发动机",
    "configuration": "构型",
    "test point": "测试点",
    "campaign": "活动",
    "ground run": "地面运行",
    "functional": "功能",
    "initial": "初始",
    "de-ice": "除冰",
    "anti-ice": "防冰",
    "windshield": "风挡",
    "wiper": "雨刷",
    "heating": "加热",
    "cooling": "冷却",
    "stability": "稳定性",
    "control": "控制",
    "load bank": "负载组",
    "ballast": "配重",
    "interior": "内饰",
    "cockpit": "驾驶舱",
    "avionics": "航电",
    "radar": "雷达",
    "transponder": "应答机",
    "communication": "通信",
    "navigation": "导航",
    "emergency": "应急",
    "hydraulic": "液压",
    "electrical": "电气",
    "pneumatic": "气动",
    "ice protection": "防冰",
    "landing gear": "起落架",
    "brake": "刹车",
    "steering": "转向",
    "nose wheel": "前轮",
    "compensator": "补偿器",
    "actuator": "作动器",
    "sensor": "传感器",
    "harness": "线束",
    "cable": "电缆",
    "panel": "面板",
    "indicator": "指示器",
    "flight hour": "飞行小时",
    "sortie": "架次",
    "data evaluation": "数据评估",
    "flight control": "飞行控制",
    "auto pilot": "自动驾驶",
    "autopilot": "自动驾驶",
    "flap": "襟翼",
    "slat": "缝翼",
    "spoiler": "扰流板",
    "elevator": "升降舵",
    "rudder": "方向舵",
    "aileron": "副翼",
    "longitudinal": "纵向",
    "lateral": "横向",
    "directional": "方向",
    "airspeed": "空速",
    "altitude": "高度",
    "instrument": "仪表",
    "approach": "进近",
    "missed approach": "复飞",
    "go-around": "复飞",
    "performance": "性能",
    "endurance": "续航时间",
    "crosswind": "侧风",
    "tailwind": "顺风",
    "headwind": "逆风",
    "icing": "结冰",
    "vibration": "振动",
    "flutter": "颤振",
    "noise": "噪声",
    "structural": "结构",
    "static": "静力",
    "dynamic": "动力",
    "electromagnetic": "电磁",
    "compatibility": "兼容性",
    "lightning": "闪电",
    "honeywell": "霍尼",
    "inertial": "惯性",
    "separator": "分离器",
    "content freeze": "内容冻结",
    "data delivery": "数据提交",
    "final stage": "最终阶段",
    "plan review": "计划审查",
    "weight and balance": "重量平衡",
    "current config": "当前构型",
    "limited": "限制",
    "post": "后",
    "risk": "风险",
    "minimizing": "最小化",
    "partial": "部分完成",
    "remaining": "剩余",
    "schedule": "计划",
    "seastar": "海星飞机",
    "ferry": "调机",
    "final": "最终",
    "fcs": "飞控系统",
    "ecs": "环控系统",
    "fqis": "燃油量指示系统",
    "adau": "ADAU",
    "dau": "DAU",
    "easa": "EASA",
    "content": "内容",
    "freeze": "冻结",
    "rl": "红标件",
    "hw": "硬件",
    "sw": "软件",
    "familiarization": "熟悉",
    "assessment": "评估",
    "cabin": "客舱",
    "interior removal": "内饰拆除",
    "certifiable": "认证",
    "lever": "手柄",
    # RMKS / Notes phrases (Col 8)
    "all test points": "所有测试点",
    "if not exchange": "如不交换",
    "any t/o tp can be used": "任意起飞测试点可用",
    "can be combined": "可合并",
    "can be performed at the end": "可在最后执行",
    "can only be planned": "仅可计划使用",
    "crosswind test results can be used": "侧风测试结果可使用",
    "data can be used": "数据可用于",
    "depends on easa availability": "取决于EASA可用性",
    "dual fms check": "双FMS检查",
    "edmo activity, parallel activity": "基地活动，并行活动",
    "engine airstart": "发动机空中起动",
    "isep rigging check": "ISEP装配检查",
    "land ops tps only": "仅陆地运行测试点",
    "no aft cg": "无后重心",
    "no aft cg limit": "无后重心限制",
    "not required for cert": "取证非必需",
    "remaining items will be performed": "剩余项目将在",
    "to be coordinated": "待协调",
    "to be planned": "待计划",
    "water ops tps only": "仅水上运行测试点",
    # Prerequisite phrases (Col 7)
    "final adau": "最终ADAU",
    "final flap": "最终襟翼",
    "ecu final sw": "环控最终软件",
    "rainy weather": "雨天",
    "pilot force gauges": "飞行员操纵力计",
    "dgps at arp": "机场基准点DGPS",
    "ruska at arp": "机场基准点RUSKA",
    "data evaluation completed": "数据评估已完成",
    "crosswind to be completed": "侧风测试待完成",
    "load bank installed": "负载组已安装",
    "generator cooling final config": "发电机冷却最终构型",
    "gcu-12 config": "GCU-12构型",
    "engine shutdown": "发动机关车",
    "glareshield design": "遮光板设计",
    "lsa tape enabled": "LSA带启用",
    "repaired curtiss wright adau delivery": "返修Curtiss Wright ADAU交付",
    "fcn lim to be extended": "FCN限制待扩展",
    "aft cg of overweight": "超重后重心",
    "safe flight": "安全飞行",
    # Reason for specific A/C (Col 14)
    "aft cg": "后重心",
    "no 6ft/sec limitation": "无6英尺/秒限制",
    "sn1004 dme is not certifiable": "SN1004 DME不可取证",
    "vmo excedence only allowed for sn1003": "VMO超速仅允许SN1003",
    # Additional Requirements (Col 15)
    "overnight": "过夜",
    "fuel imbalance": "燃油不平衡",
    "heavy fwd": "重前重心",
    "light fwd": "轻前重心",
    "aft cg 26%": "后重心26%",
    "load bank": "负载组",
    "hyd design freeze": "液压设计冻结",
    "malfunctions": "故障",
    "reservoir": "油箱",
    "relight lim to be removed": "重新点火限制待解除",
    "oei tp": "单发测试点",
    "warning check test requires": "警告检查测试需要",
    "final test": "最终测试",
    "fire": "火",
    "runway extension": "跑道延长",
    # Missing task name translations
    "engine airstart lim?": "发动机空中启动限制？",
    "engine airstart": "发动机空中启动",
    "single generator failure test": "单发电机失效测试",
    "dual generator failure test": "双发电机失效测试",
    "normal electrical power operation and bus transition tests": "正常电源运行和汇流条转换测试",
    # Complete full phrases for better translation
    "dgps at arp, ruska at arp": "基准点DGPS、RUSKA",
    "fcn lim to be extended for aft cg of overweight": "超重后重心的FCN限制待扩展",
    "gcu-12 config to allow engine shutdown": "GCU-12构型以允许发动机关车",
    "glareshield design to be changed": "遮光板设计待改变",
    "lsa tape enabled adau": "LSA带启用ADAU",
    "stall warning final tc config": "失速警告最终TC构型",
    "unusable fuel testing had to be completed": "不可用燃油测试必须已完成",
    "all test points (rl 5.2.1 on board)": "所有测试点（红标件5.2.1机上安装）",
    "data can be used for stall speed determination": "数据可用于失速速度确定",
    "depends on easa availability (week 36 is planned)": "取决于EASA可用性（计划第36周）",
    "engine airstart lim?": "发动机空中起动限制？",
    "hw sw delivery: 05.06.2026": "硬件软件交付：05.06.2026",
    "isep rigging check, parallel activity": "ISEP装配检查，并行活动",
    "night flight coordination is required": "需要夜间飞行协调",
    "one test point to be tested": "单个测试点待测试",
    "phase 1&2 with adjusted lift transducer": "阶段1和2采用调整后的升力传感器",
    "requires tower fly-by coordination": "需要塔台飞越协调",
    "spot checks to verify 26% cg": "抽查以验证26%重心",
    "tps can be used from 2ft frq": "测试点可从2FT FRQ中使用",
    "vor to stuttgart": "VOR至斯图加特",
    "t0+17": "起飞后17分钟",
    "can be performed in parallel": "可并行执行",
    "weekend days will be used": "周末将使用",
    "final test , fire and runway extersion risk": "最终测试、火灾和跑道延长风险",
    "relight lim to be removed (gcu-12)": "GCU-12重启限制待移除",
    "relight lim to be removed (gcu-12) for oei tp": "GCU-12单发测试点重启限制待移除",
    # Additional prerequisite/remarks phrases
    "32 degree outside": "32度外翼构型",
    "dgps at arp": "基准点DGPS",
    "ruska at arp": "基准点RUSKA",
    "fcn lim to be extended": "FCN限制待扩展",
    "aft cg of overweight": "超重后重心",
    "gcu-12 config": "GCU-12构型",
    "engine shutdown": "发动机关车",
    "glareshield design": "遮光板设计",
    "lsa tape enabled": "LSA带启用",
    "adau": "ADAU",
    "stall warning final": "失速警告最终配置",
    "unusable fuel testing": "不可用燃油测试",
    "rl 5.2.1": "红标件5.2.1",
    "sl 5.3": "银标件5.3",
    "all test points": "所有测试点",
    "on board": "机上安装",
    "can be performed at the end": "可在最后执行",
    "adjusting": "调整",
    "w&cg": "重量和重心",
    "can only be planned": "仅可计划",
    "sn1003": "SN1003飞机",
    "sn1004": "SN1004飞机",
    "stall speed determination": "失速速度确定",
    "depends on easa availability": "取决于EASA可用性",
    "week 36": "第36周",
    "engine airstart": "发动机空中起动",
    "isep rigging check": "ISEP装配检查",
    "parallel activity": "并行活动",
    "if not exchange": "如不交换",
    "night flight coordination": "夜间飞行协调",
    "one test point": "单个测试点",
    "requires tower fly-by": "需要塔台飞越",
    "tower fly-by coordination": "塔台飞越协调",
    "phase 1": "阶段1",
    "phase 2": "阶段2",
    "lift transducer": "升力传感器",
    "fuel imbalance": "燃油不平衡",
    "400lbs": "400磅",
    "relight lim": "重新点火限制",
    "oei tp": "单发测试点",
    "warning check test": "警告检查测试",
    "rl5.2.1": "红标件5.2.1",
    "fire and runway": "火灾和跑道",
    "runway extersion": "跑道延长",
    "current config": "当前构型",
    "current plan": "当前计划",
    "hwsw delivery": "硬件软件交付",
    "data can be used": "数据可用于",
    # More phrases from requirements
    "stall speed determination": "失速速度确定",
    "to be extended": "待扩展",
    "to be changed": "待改变",
    "to allow": "以允许",
    "to be removed": "待移除",
    "to be completed": "待完成",
    "to be used": "可用于",
    "to verify": "用于验证",
    "with adjusted": "调整后",
    "can be performed in parallel": "可并行执行",
    "spot checks": "抽查",
    "weekend days": "周末",
    "will be used": "将使用",
    "tps can be": "测试点可以",
    "from 2ft frq": "来自2FT FRQ",
    "vor to": "VOR至",
    "t0+17": "起飞后17分钟",
    "rgps": "DGPS",
    "atp": "机场",
    "fcn": "FCN",
    "lim": "限制",
    "aft cg of": "后重心",
    "gcu": "GCU",
    "config to allow": "构型以允许",
    "tc config": "TC构型",
    "testing had to be": "测试必须已",
    "unusable fuel": "不可用燃油",
    "all test points": "所有测试点",
    "rl 5.2.1 on board": "红标件5.2.1机上安装",
    "rl5.2.1": "红标件5.2.1",
    "depends on easa": "取决于EASA",
    "week 36 is planned": "计划第36周",
    "engine airstart lim": "发动机空中起动限制",
    "isep rigging": "ISEP装配",
    "night flight": "夜间飞行",
    "one test point": "单个测试点",
    "to be tested": "待测试",
    "phase 1&2": "阶段1和2",
    "lift transducer": "升力传感器",
    "requires tower": "需要塔台",
    "fly-by": "飞越",
    "coordination": "协调",
    "26% cg": "26%重心",
    "relight": "重新点火",
    "final test": "最终测试",
    "fire and runway": "火灾和跑道",
    "extersion risk": "延长风险",
    "oei tp": "单发测试点",

}


# ── Fixed mappings for specific columns ──
PAPERWORK_STATUS_MAP = {
    "in work": "进行中",
    "n/a": "不适用",
    "not started": "未开始",
    "released": "已发布",
    "signature loop": "签字中",
    "waiting ftp": "等待FTP",
}


def parse_glossary(md_path):
    """Parse markdown glossary. Returns dict: English(lowercase) -> Chinese.
    Now supports multiple columns (Task Name, Prerequisite, Remarks, etc.)."""
    glossary = {}
    if not os.path.exists(md_path):
        return glossary

    current_section = None
    with open(md_path, "r", encoding="utf-8") as f:
        for line in f:
            stripped = line.strip()

            # Detect section headers
            if stripped.startswith("## ") and "Translations" in stripped:
                current_section = stripped.replace("## ", "").replace(" Translations", "")
                continue

            # Skip markdown separators and empty lines
            if not stripped or "|---|---|" in stripped or (stripped.startswith("|") and "EN" in stripped):
                continue

            # Parse table rows
            if stripped.startswith("|"):
                parts = [p.strip() for p in stripped.split("|")]
                if len(parts) >= 4:
                    en = parts[1]
                    cn = parts[2]
                    if en and cn and en.lower() != "en" and en != "(none)":
                        # Store with lowercase key for case-insensitive matching
                        # If multiple translations for same EN, keep first one
                        en_lower = en.lower()
                        if en_lower not in glossary:
                            glossary[en_lower] = cn

    return glossary


def strip_frq_prefix(name):
    """Remove FRQ prefix from task name."""
    m = re.match(r'^(?:FRQ[-_]\d+(?:[-_]\d+(?:\.\d+)?(?:[-_]\d+)?)?\s*)(.*)', name)
    return m.group(1).strip() if m else name


def strip_indent(name):
    return name.lstrip()


def look_up_task_name(name, glossary):
    """Full lookup pipeline for task names."""
    stripped = strip_indent(name).strip()
    if not stripped:
        return name, "empty"
    has_chinese = bool(re.search(r'[一-鿿]', stripped))
    has_ascii = bool(re.search(r'[A-Za-z]{2,}', stripped))
    if has_chinese and not has_ascii:
        return stripped, "already-cn"

    lower = stripped.lower()

    # 0. Try exact match in BUILTIN_DICT first (prioritize for task names)
    if lower in BUILTIN_DICT:
        return BUILTIN_DICT[lower], "builtin-exact"

    # 1. Try exact match in glossary (case-insensitive)
    if lower in glossary:
        return glossary[lower], "exact"

    # 2. Try without punctuation in BUILTIN_DICT
    lower_no_punct = re.sub(r'[?!.,;:\'"()]', '', lower).strip()
    if lower_no_punct in BUILTIN_DICT:
        return BUILTIN_DICT[lower_no_punct], "builtin-no-punct"

    # 3. Case-insensitive search through glossary keys
    for key in glossary.keys():
        if key == lower:
            return glossary[key], "exact-ci"

    # 4. FRQ prefix stripping
    core = strip_frq_prefix(stripped)
    if core != stripped:
        core_lower = core.lower()
        if core_lower in glossary:
            return glossary[core_lower], "frq-exact"
        if core_lower in BUILTIN_DICT:
            return BUILTIN_DICT[core_lower], "frq-builtin"

    # 5. Fuzzy matching
    keys = list(glossary.keys())
    matches = difflib.get_close_matches(lower, keys, n=1, cutoff=0.85)
    if matches:
        return glossary[matches[0]], "fuzzy"

    if core != stripped:
        core_lower = core.lower()
        matches = difflib.get_close_matches(core_lower, keys, n=1, cutoff=0.85)
        if matches:
            return glossary[matches[0]], "fuzzy-frq"

    # 6. Phrase translate (word-by-word)
    trans = phrase_translate(stripped)
    if trans:
        return trans, "phrase-translate"

    # 7. Keyword matching
    stopwords = {'the','a','an','and','or','for','of','to','in','on','at',
                 'with','by','is','are','be','was','were','been',
                 'test','testing','system','from','until','this','that'}
    core_words = set(re.findall(r'[A-Za-z]+', stripped.lower())) - stopwords
    if core_words:
        best_key, best_score = None, 0
        for k in keys:
            kw = set(re.findall(r'[A-Za-z]+', k.lower())) - stopwords
            if not kw: continue
            inter, union = core_words & kw, core_words | kw
            score = len(inter)/len(union) if union else 0
            if score > best_score and score >= 0.4:
                best_score, best_key = score, k
        if best_key:
            return glossary[best_key], "keyword(%.2f)" % best_score

    return stripped, "not-found"


def phrase_translate(text):
    """Translate a phrase using the built-in dictionary (word-by-word fallback)."""
    lower = text.lower().strip()
    if lower in BUILTIN_DICT:
        return BUILTIN_DICT[lower]

    # Try without punctuation
    lower_no_punct = re.sub(r'[?!.,;:\'"()]', '', lower).strip()
    if lower_no_punct in BUILTIN_DICT:
        return BUILTIN_DICT[lower_no_punct]

    # Word-by-word translation
    parts = re.split(r'(\s+|[-/]|[?!.,;:\'"()])', text)
    translated, changed = [], False
    for part in parts:
        plow = part.lower().strip()
        if not plow or re.match(r'^[\s\-/?!.,;:\'"()]+$', part):
            translated.append(part)
        elif re.match(r'^(FRQ|SN|MSN|WOT|EDMO|EDXW|LEPO|[A-Z]{2,6})$', part, re.IGNORECASE):
            translated.append(part)
        elif re.match(r'^[\d.]+$', part):
            translated.append(part)
        else:
            found = BUILTIN_DICT.get(plow)
            if found:
                translated.append(found); changed = True
            else:
                translated.append(part)
    return "".join(translated) if changed else None


def translate_cell_text(text, is_task_name, glossary):
    """Translate a single cell value. Uses different strategies per column type.
    For non-task-names, prioritizes glossary lookup over builtin dict."""
    if not text or not text.strip():
        return text, False

    stripped = text.strip()
    has_cn = bool(re.search(r'[一-鿿]', stripped))
    if has_cn:
        return stripped, False  # Already Chinese

    if is_task_name:
        result, _ = look_up_task_name(stripped, glossary)
        if result != stripped:
            return result, True
    else:
        # For other columns: PRIORITIZE glossary lookups
        lower = stripped.lower()

        # 1. Try exact match in glossary (case-insensitive)
        if lower in glossary:
            return glossary[lower], True

        # 2. Try matching full phrase from glossary with fuzzy matching
        glossary_keys = list(glossary.keys())
        matches = difflib.get_close_matches(lower, glossary_keys, n=1, cutoff=0.85)
        if matches:
            matched_key = matches[0]
            if matched_key in glossary:
                return glossary[matched_key], True

        # 3. Try phrase lookup in builtin dict
        if lower in BUILTIN_DICT:
            return BUILTIN_DICT[lower], True

        # 4. Try paperwork status map
        if lower in PAPERWORK_STATUS_MAP:
            return PAPERWORK_STATUS_MAP[lower], True

        # 5. Word-by-word phrase translation
        trans = phrase_translate(stripped)
        if trans and trans != stripped:
            return trans, True

    return stripped, False


def translate_all_columns(ws, glossary):
    """Translate all text columns in the Excel sheet."""
    max_row = ws.max_row

    # Columns to translate: (col_index, label, is_task_name)
    translate_cols = [
        (3, "Task Name", True),
        (7, "Prerequisite", False),
        (8, "RMKS / Notes", False),
        (9, "Paperwork status", False),
        (14, "Reason for specific A/C", False),
        (15, "Additional Requirements", False),
    ]

    col_stats = {col: {"translated": 0, "total": 0} for col, _, _ in translate_cols}

    for row in range(2, max_row + 1):
        for col, label, is_task_name in translate_cols:
            cell = ws.cell(row, col)
            value = str(cell.value or "").strip()
            if not value or value == "None":
                continue

            col_stats[col]["total"] += 1

            translated, changed = translate_cell_text(value, is_task_name, glossary)
            if not changed:
                continue

            col_stats[col]["translated"] += 1

            # Preserve cell formatting
            orig_font = cell.font
            orig_alignment = cell.alignment
            cell.value = translated
            cell.font = Font(
                name=orig_font.name or "Calibri",
                size=orig_font.size or 11,
                bold=orig_font.bold or False,
                color=orig_font.color,
            )
            cell.alignment = Alignment(
                horizontal=orig_alignment.horizontal or "left",
                vertical=orig_alignment.vertical or "center",
                wrap_text=True if label in ("RMKS / Notes", "Prerequisite", "Additional Requirements") else (
                    orig_alignment.wrap_text if orig_alignment else False),
            )

    # Print stats
    print()
    for col, label, _ in translate_cols:
        stats = col_stats[col]
        print("  Col %d (%s): translated %d/%d" % (col, label, stats["translated"], stats["total"]))

    return sum(s["translated"] for s in col_stats.values())


def main():
    parser = argparse.ArgumentParser(description="Translate English text to Chinese")
    parser.add_argument("--input", required=True, help="Input .xlsx (Phase 2 output)")
    parser.add_argument("--glossary", default="references/术语对照.md", help="Glossary .md path")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print("Error: Input file not found: %s" % args.input, file=sys.stderr)
        sys.exit(1)

    print("Loading glossary: %s" % args.glossary)
    glossary = parse_glossary(args.glossary)
    print("  Loaded %d glossary entries" % len(glossary))

    print("Loading workbook: %s" % args.input)
    wb = openpyxl.load_workbook(args.input)
    ws = wb.active
    print("  Sheet: '%s', rows: %d" % (ws.title, ws.max_row))

    print("Translating...")
    count = translate_all_columns(ws, glossary)

    print("\nSaving: %s" % args.output)
    wb.save(args.output)
    print("  Total cells translated: %d" % count)
    print("Done!")


if __name__ == "__main__":
    main()
