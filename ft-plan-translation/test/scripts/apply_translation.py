#!/usr/bin/env python3
"""Apply Chinese translation mapping to one week of the flight-test planning
workbook.  Edits the .xlsm via direct XML surgery (zip-level) so VBA macros,
images, and data validation survive byte-for-byte.

Usage:
    python apply_translation.py <in.xlsm> <sheet: MSN1003|MSN1004> <week> <mapping.json> <out.xlsm>

Features (v3):
  - Font consistency: Chinese runs inherit the English run's colour, font name,
    and font size (capped at 9 pt); bold is stripped.
    Colour resolution order: explicit run rPr → CF DXF font colour (theme/indexed
    resolved via theme1.xml / indexedColors palette) → cell default font colour
    (RGB → indexed → theme). All three colour sources (RGB, theme, indexed) are
    now correctly resolved.
  - CF fill baking: reads ``cellIs equal`` CF rules, evaluates which DXF rule
    matches each cell's original EN text, then bakes the DXF fill into a new
    cellXf style.  The ``_ensure_cellxf`` helper uses a split-on-boundary
    approach to handle both self-closing ``<xf …/>`` and child-bearing
    ``<xf>…</xf>`` entries, avoiding the earlier regex bug that miscounted
    entries and silently skipped fill baking.
  - Row hiding: after translation ALL rows except header rows 1-3 and the
    ENTIRE 10-row week block (week_row … week_row+9) are hidden.  For
    multi-week chains the caller MUST do a final visibility pass to unhide the
    rows of ALL translated weeks — the script only keeps its own week visible.
  - Multi-week safety: ``_hide_rows`` now *removes* ``hidden="1"`` from rows
    that should be visible, so chained runs don't leave stale hidden state.
"""

import copy
import json
import re
import sys
import zipfile
from xml.sax.saxutils import escape

# ---------------------------------------------------------------------------
# shared constants
# ---------------------------------------------------------------------------
M_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
FH_RE = re.compile(r"^\s*\d+\s*FH Inspection\b.*$", re.I)

WEEK_ROW0 = 5
WEEK_STRIDE = 10
BLOCK_ROWS = 9                 # rows after week_row that belong to the block
COL_MIN, COL_MAX = 2, 11       # B..K

# CF rule: ``cellIs equal`` whose formula is a double-quoted literal like
# ``"Flight Test"`` or ``"Maintenance"``.
CF_FORMULA_RE = re.compile(r'^"([^"]*)"$')

# ---------------------------------------------------------------------------
# XML helpers
# ---------------------------------------------------------------------------
def _sheet_xml_name(z, sheet):
    wb = z.read("xl/workbook.xml").decode()
    rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb))[sheet]
    rels = z.read("xl/_rels/workbook.xml.rels").decode()
    tgt = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))[rid]
    return "xl/" + tgt.lstrip("/")


def _unescape(s):
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'").replace("&amp;", "&"))


def _col_number(col_ref):
    n = 0
    for ch in col_ref:
        n = n * 26 + ord(ch.upper()) - 64
    return n


# ---------------------------------------------------------------------------
# shared-string parsing / building
# ---------------------------------------------------------------------------
def parse_si(si_xml):
    """Return list of (text, rPr_xml_or_None)."""
    runs = []
    if "<r>" not in si_xml and "<r " not in si_xml:
        m = re.search(r"<t[^>]*>(.*?)</t>", si_xml, re.S)
        if m:
            runs.append((_unescape(m.group(1)), None))
        return runs
    for rm in re.finditer(r"<r>(.*?)</r>", si_xml, re.S):
        body = rm.group(1)
        pr = re.search(r"(<rPr>.*?</rPr>)", body, re.S)
        tm = re.search(r"<t[^>]*>(.*?)</t>", body, re.S)
        runs.append((_unescape(tm.group(1)) if tm else "", pr.group(1) if pr else None))
    return runs


def runs_to_lines(runs):
    """Split runs at newlines -> list of lines, each list of (text, rPr)."""
    lines, cur = [], []
    for text, pr in runs:
        parts = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        for i, part in enumerate(parts):
            if i > 0:
                lines.append(cur)
                cur = []
            if part:
                cur.append((part, pr))
    lines.append(cur)
    return lines


def _strip_bold(rpr_xml):
    if not rpr_xml:
        return None
    return re.sub(r"<b\s*/>", "", rpr_xml)


def _cap_sz(rpr_xml, max_sz=9):
    """Cap <sz val="N"/> so N ≤ max_sz.  Return (rpr_xml, changed)."""
    if not rpr_xml:
        return rpr_xml, False
    m = re.search(r'<sz\s+val="([\d.]+)"', rpr_xml)
    if not m:
        return rpr_xml, False
    cur = float(m.group(1))
    if cur <= max_sz:
        return rpr_xml, False
    new = re.sub(r'(<sz\s+val=")[\d.]+(")', rf'\g<1>{max_sz:g}\g<2>', rpr_xml, count=1)
    return new, True


def _merge_sz(rpr_xml, target_sz):
    """Ensure <sz> is present with the given value (if not already present)."""
    if not rpr_xml:
        return f'<rPr><sz val="{target_sz:g}"/></rPr>'
    if '<sz ' in rpr_xml:
        return rpr_xml
    return rpr_xml.replace('<rPr>', f'<rPr><sz val="{target_sz:g}"/>', 1)


def _cn_rpr(line_runs, default_rpr, cf_font_color=None, max_size=9):
    """Build rPr string for the Chinese run from the line's first styled run.

    Priority: explicit run rPr → CF font colour → cell default.
    Bold is always stripped and font size capped at *max_size* pts.

    Font name, size, and family are inherited from *default_rpr* when the
    source run rPr doesn't provide them, so Chinese text always has explicit
    formatting that matches the English cell font.
    """
    # Find the first run that carries an rPr
    src_rpr = None
    for _, pr in line_runs:
        if pr:
            src_rpr = pr
            break
    if not src_rpr:
        src_rpr = default_rpr

    rpr = _strip_bold(src_rpr)

    # Inherit font name from default_rpr if source rPr doesn't specify one
    if default_rpr and '<rFont' not in (rpr or ''):
        m = re.search(r'<rFont[^>]*/>', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0), 1)
            else:
                rpr = f'<rPr>{m.group(0)}</rPr>'

    # Inherit font family from default_rpr if missing
    if default_rpr and '<family' not in (rpr or ''):
        m = re.search(r'<family\s+val="[^"]*"', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0) + '/>', 1)
            else:
                rpr = f'<rPr>{m.group(0)}/></rPr>'

    # Inherit font size from default_rpr if missing
    if default_rpr and '<sz ' not in (rpr or ''):
        m = re.search(r'<sz\s+val="[\d.]+"', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0) + '/>', 1)
            else:
                rpr = f'<rPr>{m.group(0)}/></rPr>'

    # If the run has no explicit colour, inject the CF colour (if any).
    # Match '<color' without trailing space to also catch self-closing
    # tags like '<color/>'.
    if cf_font_color and ('<color' not in (rpr or '')):
        if rpr:
            rpr = rpr.replace('<rPr>', f'<rPr><color rgb="{cf_font_color}"/>', 1)
        else:
            rpr = f'<rPr><color rgb="{cf_font_color}"/></rPr>'

    rpr, _ = _cap_sz(rpr, max_size)
    return rpr


def build_si(lines_out):
    """lines_out: list of lines, each list of (text, rPr).  Return <si> XML."""
    parts = ["<si>"]
    flat = []
    for i, line in enumerate(lines_out):
        if i > 0:
            flat.append(("\r\n", line[0][1] if line else None))
        flat.extend(line if line else [])
    # merge adjacent runs with identical rPr
    merged = []
    for text, pr in flat:
        if merged and merged[-1][1] == pr:
            merged[-1] = (merged[-1][0] + text, pr)
        else:
            merged.append((text, pr))
    for text, pr in merged:
        t = f'<t xml:space="preserve">{escape(text)}</t>'
        parts.append(f"<r>{pr or ''}{t}</r>")
    parts.append("</si>")
    return "".join(parts)


# ---------------------------------------------------------------------------
# conditional-formatting engine
# ---------------------------------------------------------------------------
def _parse_dxfs(styles_xml):
    """Return list of dicts {fill_xml, font_color_rgb, font_theme, bold, font_xml}.

    Per OOXML, font colour can be one of:
      <color rgb="FFNNNNNN"/>    -- explicit ARGB
      <color theme="N"/>          -- theme colour
      <color indexed="N"/>        -- indexed colour
    If the DXF has no <font> element, font keys are None.
    """
    dxfs = re.findall(r"<dxf>(.*?)</dxf>", styles_xml, re.S)
    parsed = []
    for d in dxfs:
        entry = {"fill_xml": None, "font_color_rgb": None,
                 "font_theme": None, "font_color_indexed": None,
                 "bold": False, "font_xml": None}
        fill_m = re.search(r"<fill>(.*?)</fill>", d, re.S)
        if fill_m:
            entry["fill_xml"] = fill_m.group(1).strip()
            if not entry["fill_xml"]:
                entry["fill_xml"] = None
        font_m = re.search(r"<font>(.*?)</font>", d, re.S)
        if font_m:
            font_body = font_m.group(1)
            entry["font_xml"] = font_body
            entry["bold"] = "<b/>" in font_body
            color_m = re.search(r'<color\s+(.*?)/?>', font_body)
            if color_m:
                cattrs = color_m.group(1)
                crgb = re.search(r'rgb="([^"]*)"', cattrs)
                ctheme = re.search(r'theme="([^"]*)"', cattrs)
                cidx = re.search(r'indexed="([^"]*)"', cattrs)
                if crgb:
                    entry["font_color_rgb"] = crgb.group(1)
                if ctheme:
                    entry["font_theme"] = ctheme.group(1)
                if cidx:
                    entry["font_color_indexed"] = cidx.group(1)
        parsed.append(entry)
    return parsed


def _parse_cf_rules(sheet_xml):
    """Return list of (sqref_parts, list_of_rule_dicts).

    Each rule dict: {type, operator, formula_text, dxfId, priority}
    Only ``cellIs equal`` rules with a literal ``"text"`` formula are kept.
    """
    blocks = re.findall(
        r'<conditionalFormatting\s+sqref="([^"]*)"\s*>(.*?)</conditionalFormatting>',
        sheet_xml, re.S)
    rules = []
    for sqref, body in blocks:
        parts = sqref.split()
        rule_list = []
        for attrs_str, rule_body in re.findall(r"<cfRule(.*?)>(.*?)</cfRule>", body, re.S):
            a = dict(re.findall(r'(\w+)="([^"]*)"', attrs_str))
            if a.get("type") != "cellIs" or a.get("operator") != "equal":
                continue
            f_m = re.search(r"<formula>(.*?)</formula>", rule_body)
            if not f_m:
                continue
            fm = CF_FORMULA_RE.match(f_m.group(1).strip())
            if not fm:
                continue
            rule_list.append({
                "text": fm.group(1),
                "dxfId": int(a.get("dxfId", "0")),
                "priority": int(a.get("priority", "0")),
            })
        if rule_list:
            rules.append((parts, rule_list))
    return rules


def _cf_match(cell_col, cell_row, cell_text, cf_rules):
    """Return the highest-priority matching DXF id, or None."""
    best_pri = -1
    best_dxf = None
    col_ref = ""
    # Build column letters for sqref matching
    n = cell_col
    while n > 0:
        n, r = divmod(n - 1, 26)
        col_ref = chr(65 + r) + col_ref

    for sqref_parts, rule_list in cf_rules:
        # Check if cell is in any sqref part
        covered = False
        for part in sqref_parts:
            m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", part)
            if not m:
                continue
            c1, r1 = _col_number(m.group(1)), int(m.group(2))
            c2 = _col_number(m.group(3)) if m.group(3) else c1
            r2 = int(m.group(4)) if m.group(4) else r1
            if c1 <= cell_col <= c2 and r1 <= cell_row <= r2:
                covered = True
                break
        if not covered:
            continue

        for rule in rule_list:
            if rule["text"] == cell_text and rule["priority"] > best_pri:
                best_pri = rule["priority"]
                best_dxf = rule["dxfId"]
    return best_dxf


# ---------------------------------------------------------------------------
# theme colour resolution
# ---------------------------------------------------------------------------
def _resolve_theme_colors(z):
    """Read theme1.xml from the zip and return dict {theme_idx: "RRGGBB"}."""
    try:
        theme_xml = z.read("xl/theme/theme1.xml").decode()
    except (KeyError, zipfile.BadZipFile):
        return {}

    scheme = re.search(r"<a:clrScheme[^>]*>(.*?)</a:clrScheme>", theme_xml, re.S)
    if not scheme:
        return {}

    theme_map = {
        "dk1": 1, "lt1": 0, "dk2": 3, "lt2": 2,
        "accent1": 4, "accent2": 5, "accent3": 6,
        "accent4": 7, "accent5": 8, "accent6": 9,
    }
    colors = {}
    for name, idx in theme_map.items():
        m = re.search(
            rf"<a:{name}>\s*<a:srgbClr\s+val=\"([^\"]*)\"",
            scheme.group(1), re.S,
        )
        if not m:
            m = re.search(
                rf"<a:{name}>\s*<a:sysClr\s+lastClr=\"([^\"]*)\"",
                scheme.group(1), re.S,
            )
        if m:
            colors[idx] = m.group(1).upper()
    return colors


# ---------------------------------------------------------------------------
# indexed colour palette
# ---------------------------------------------------------------------------
def _resolve_indexed_colors(styles_xml):
    """Parse <indexedColors> from styles.xml, return {idx: 'RRGGBB'}."""
    colors = {}
    palette = re.search(r"<indexedColors>(.*?)</indexedColors>", styles_xml, re.S)
    if not palette:
        return colors
    entries = re.findall(r"<rgbColor[^>]*rgb=\"([^\"]+)\"", palette.group(1))
    for i, rgb in enumerate(entries):
        colors[i] = rgb.upper()
    return colors


# ---------------------------------------------------------------------------
# fill / cellXf management
# ---------------------------------------------------------------------------
def _ensure_fill(styles_xml, fill_xml):
    """Append a fill to styles.xml if not already present; return (new_xml, fill_idx)."""
    fills_section = re.search(r"(<fills\s+count=\")(\d+)(\">)", styles_xml)
    if not fills_section:
        raise RuntimeError("Cannot locate <fills> in styles.xml")
    count = int(fills_section.group(2))
    # Simple check: does this exact fill XML already exist?
    existing = re.findall(r"<fill>(.*?)</fill>", styles_xml, re.S)
    for i, f in enumerate(existing):
        if f.strip() == fill_xml.strip():
            return styles_xml, i

    # Append
    new_fill = f"<fill>{fill_xml}</fill>"
    new_count = count + 1
    styles_xml = styles_xml.replace(
        fills_section.group(0),
        f'<fills count="{new_count}">',
        1)
    idx = styles_xml.rfind("</fills>")
    styles_xml = styles_xml[:idx] + new_fill + styles_xml[idx:]
    return styles_xml, count


def _build_fill_xml(rgb_color):
    """Build a solid patternFill with the given ARGB background colour.

    Typical CF DXF fill: <patternFill><bgColor rgb="FFFFC000"/></patternFill>
    We convert this to a native cell fill:
      <fill><patternFill patternType="solid"><fgColor rgb="..."/><bgColor indexed="64"/></patternFill></fill>
    """
    return (
        '<patternFill patternType="solid">'
        f'<fgColor rgb="{rgb_color}"/>'
        '<bgColor indexed="64"/>'
        '</patternFill>'
    )


def _build_fill_direct(dxf_fill_xml):
    """Try to extract a solid fill colour from the DXF fill XML.

    DXF fills typically use the bgColor to specify the background.
    We extract the rgb/indexed/theme colour and build a proper cell fill.
    Returns (fill_xml, rgb_or_None).
    """
    bg_m = re.search(r'<bgColor\s*(.*?)/?>', dxf_fill_xml)
    if not bg_m:
        return None, None
    attrs = bg_m.group(1)
    rgb_m = re.search(r'rgb="([^"]*)"', attrs)
    if rgb_m:
        rgb = rgb_m.group(1)
        return _build_fill_xml(rgb), rgb
    theme_m = re.search(r'theme="([^"]*)"', attrs)
    if theme_m:
        # For theme colours we preserve the DXF fill as-is but wrapped in a
        # <fill> suitable for cellXf.  This is a best-effort approximation.
        tint_m = re.search(r'tint="([^"]*)"', attrs)
        tint = f' tint="{tint_m.group(1)}"' if tint_m else ""
        return (f'<patternFill patternType="solid">'
                f'<fgColor theme="{theme_m.group(1)}"{tint}/>'
                f'<bgColor indexed="64"/>'
                f'</patternFill>', None)
    indexed_m = re.search(r'indexed="([^"]*)"', attrs)
    if indexed_m:
        return (f'<patternFill patternType="solid">'
                f'<fgColor indexed="{indexed_m.group(1)}"/>'
                f'<bgColor indexed="64"/>'
                f'</patternFill>', None)
    return None, None


def _ensure_cellxf(styles_xml, original_xf_id, fill_idx, fill_count_before):
    """Create a new cellXf that clones *original_xf_id* but with *fill_idx*.

    Returns (styles_xml, new_xf_id).
    """
    xf_section = re.search(r"(<cellXfs\s+count=\")(\d+)(\">)", styles_xml)
    xf_count = int(xf_section.group(2))
    xfs_block = re.search(r"<cellXfs[^>]*>(.*?)</cellXfs>", styles_xml, re.S)

    # Split on <xf boundaries to capture all entries (self-closing or with children)
    parts = re.split(r"(?=<xf\b)", xfs_block.group(1))
    xf_entries_raw = [p.strip() for p in parts if p.strip().startswith("<xf")]

    if original_xf_id >= len(xf_entries_raw):
        return styles_xml, original_xf_id

    orig = xf_entries_raw[original_xf_id]
    # Replace or add fillId
    if 'fillId=' in orig:
        new_xf = re.sub(r'(fillId=\")\d+(\")', rf'\g<1>{fill_idx}\g<2>', orig, count=1)
    else:
        # Shouldn't happen; inject
        new_xf = orig.replace('<xf ', f'<xf fillId="{fill_idx}" ', 1)

    new_xml = new_xf

    # Increment cellXfs count
    new_count = xf_count + 1
    styles_xml = styles_xml.replace(
        xf_section.group(0),
        f'<cellXfs count="{new_count}">',
        1)

    idx = styles_xml.rfind("</cellXfs>")
    styles_xml = styles_xml[:idx] + new_xml + styles_xml[idx:]
    return styles_xml, xf_count  # new xf id = old count (0‑based)


# ---------------------------------------------------------------------------
# cell-level font inspection (openpyxl, read-only)
# ---------------------------------------------------------------------------
def cell_default_rpr(cell_font, cf_font_color=None, max_size=9, theme_colors=None, indexed_colors=None):
    """Build an rPr from an openpyxl Font, optionally overriding colour from CF.

    Font size is capped at *max_size* (9 pt).
    """
    bits = []
    sz = cell_font.size if cell_font.size else 9
    sz = min(sz, max_size)
    bits.append(f'<sz val="{sz:g}"/>')

    color_src = cf_font_color  # CF colour wins over base cell colour
    if not color_src and cell_font.color:
        # Try explicit RGB (openpyxl stores as a special descriptor)
        rgb_val = getattr(cell_font.color, "rgb", None)
        # openpyxl returns '00000000' for "no colour set"; anything else is a real ARGB
        if rgb_val and isinstance(rgb_val, str) and len(rgb_val) >= 8 and rgb_val != '00000000':
            color_src = rgb_val
        # Fall back to indexed colour
        if not color_src and getattr(cell_font.color, "indexed", None) is not None:
            if indexed_colors:
                color_src = indexed_colors.get(cell_font.color.indexed)
        # Fall back to theme colour
        if not color_src and getattr(cell_font.color, "theme", None) is not None:
            if theme_colors:
                color_src = theme_colors.get(cell_font.color.theme)
    if color_src:
        bits.append(f'<color rgb="{color_src}"/>')

    if cell_font.name:
        bits.append(f'<rFont val="{cell_font.name}"/>')
        bits.append('<family val="2"/>')
    return "<rPr>" + "".join(bits) + "</rPr>" if bits else None


# ---------------------------------------------------------------------------
# line translation logic
# ---------------------------------------------------------------------------
def translate_lines(lines, mapping, default_rpr, cf_font_color=None, cn_max_sz=9):
    """Insert Chinese translations; return (new_lines, changed?)."""
    out, changed, i = [], False, 0
    while i < len(lines):
        text_i = "".join(t for t, _ in lines[i])
        if FH_RE.match(text_i):
            j, nums, suffix = i, [], ""
            while j < len(lines):
                tj = "".join(t for t, _ in lines[j])
                if not FH_RE.match(tj):
                    break
                nums.append(re.match(r"\s*(\d+)", tj).group(1))
                sm = re.match(r"^\s*\d+\s*FH Inspection\b\s*(.*)$", tj, re.I)
                suffix = sm.group(1).strip()
                out.append(lines[j])
                j += 1
            key = "/".join(nums) + "FH Inspection" + (f" {suffix}" if suffix else "")
            cn = mapping.get(key, "").strip()
            if cn:
                cn_pr = _cn_rpr(lines[j - 1], default_rpr, cf_font_color, cn_max_sz)
                out.append([(cn, cn_pr)])
                changed = True
            i = j
        else:
            out.append(lines[i])
            cn = mapping.get(text_i.strip(), "").strip()
            if cn:
                cn_pr = _cn_rpr(lines[i], default_rpr, cf_font_color, cn_max_sz)
                out.append([(cn, cn_pr)])
                changed = True
            i += 1
    return out, changed


# ---------------------------------------------------------------------------
# row hiding
# ---------------------------------------------------------------------------
def _hide_rows(sheet_xml, week_row):
    """Hide ALL rows except header rows 1-3 and the ENTIRE 10-row block of the
    target week (week_row through week_row+9: CW, AM, AM detail, overflow, PM,
    PM detail, overflow, RMK, overflow, blank separator).
    """
    visible_rows = {1, 2, 3} | set(range(week_row, week_row + 10))

    def _maybe_hide(m):
        r = int(m.group(1))
        if r not in visible_rows:
            if 'hidden="1"' not in m.group(0):
                # Insert hidden="1" into the row tag
                return re.sub(r'(<row\b[^>]*)>', r'\1 hidden="1">', m.group(0), count=1)
        else:
            # Remove hidden="1" if present (for chained multi-week runs)
            if 'hidden="1"' in m.group(0):
                return m.group(0).replace(' hidden="1"', '').replace(' hidden="1"', '')
        return m.group(0)

    sheet_xml = re.sub(r'<row\b[^>]*r="(\d+)"[^>]*>', _maybe_hide, sheet_xml)
    return sheet_xml


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    src, sheet, week, map_path, dst = (sys.argv[1], sys.argv[2],
                                       int(sys.argv[3]), sys.argv[4], sys.argv[5])
    data = json.load(open(map_path, encoding="utf-8-sig"))
    mapping = data["to_translate"]
    cells = data["cells"]
    week_row = data["week_row"]

    # ---------- openpyxl pass (READ-ONLY, never save) ----------
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(src, read_only=False, keep_vba=True)
    ws = wb[sheet]

    # ---------- read zip internals ----------
    zin = zipfile.ZipFile(src)
    sheet_name = _sheet_xml_name(zin, sheet)
    sst_xml = zin.read("xl/sharedStrings.xml").decode()
    sheet_xml = zin.read(sheet_name).decode()
    styles_xml = zin.read("xl/styles.xml").decode()

    si_bodies = re.findall(r"<si>.*?</si>", sst_xml, re.S)
    n_si = len(si_bodies)

    # ---------- conditional formatting ----------
    cf_rules = _parse_cf_rules(sheet_xml)
    dxfs = _parse_dxfs(styles_xml)
    theme_colors = _resolve_theme_colors(zin)  # resolve theme idx -> RGB
    indexed_colors = _resolve_indexed_colors(styles_xml)  # resolve indexed idx -> RGB

    new_sis, si_cache = [], {}
    cell_new_index = {}           # ref -> new SI index
    cell_new_style = {}           # ref -> new cellXf id (when CF fill needs baking)
    new_xf_fill_map = {}          # (original_xfId, fill_idx) -> new_xfId

    styles_modified = False
    fill_additions = {}           # dxf_fill_xml -> fill_idx
    original_fill_count = int(re.search(r'<fills\s+count="(\d+)"', styles_xml).group(1))

    # ---------- process each cell ----------
    for ref, info in cells.items():
        si_idx = info["si"]
        runs = parse_si(si_bodies[si_idx])
        lines = runs_to_lines(runs)

        # Determine cell position
        col_m = re.match(r"([A-Z]+)(\d+)", ref)
        col_num = _col_number(col_m.group(1))
        row_num = int(col_m.group(2))

        # Full cell text (for CF matching)
        cell_text = "".join(r[0] for r in runs)

        # Evaluate CF
        dxf_id = _cf_match(col_num, row_num, cell_text, cf_rules)
        cf_fill_xml = None
        cf_font_color = None
        if dxf_id is not None and dxf_id < len(dxfs):
            dxf = dxfs[dxf_id]
            if dxf["fill_xml"]:
                cf_fill_xml = dxf["fill_xml"]
            cf_font_color = dxf["font_color_rgb"]
            # Resolve theme colours if no explicit RGB
            if not cf_font_color and dxf["font_theme"] is not None:
                cf_font_color = theme_colors.get(int(dxf["font_theme"]))
            if not cf_font_color and dxf["font_color_indexed"] is not None:
                cf_font_color = indexed_colors.get(int(dxf["font_color_indexed"]))

        # ---------- inject CF font colour into English runs ----------
        # After translation the cell text changes (bilingual), so the CF
        # "cellIs equal" rule no longer matches.  English runs that relied
        # on CF for their font colour would lose it.  Pre-inject the colour
        # so both English and Chinese runs keep the correct colour.
        if cf_font_color:
            for line in lines:
                for i, (text, rpr) in enumerate(line):
                    if '<color' not in (rpr or ''):
                        if rpr:
                            line[i] = (text, rpr.replace(
                                '<rPr>', f'<rPr><color rgb="{cf_font_color}"/>', 1))
                        else:
                            line[i] = (text,
                                       f'<rPr><color rgb="{cf_font_color}"/></rPr>')

        # Default rPr
        default = cell_default_rpr(ws[ref].font, cf_font_color, max_size=9, theme_colors=theme_colors, indexed_colors=indexed_colors)

        new_lines, changed = translate_lines(lines, mapping, default, cf_font_color, cn_max_sz=9)
        if not changed:
            continue

        key = (si_idx, default)
        if key not in si_cache:
            # If CF provides an explicit fill, also apply it to the new SI
            si_cache[key] = n_si + len(new_sis)
            new_sis.append(build_si(new_lines))
        cell_new_index[ref] = si_cache[key]

        # ---------- CF fill baking ----------
        if cf_fill_xml:
            fill_xml_for_cell, _rgb = _build_fill_direct(cf_fill_xml)
            if fill_xml_for_cell:
                # Ensure fill exists
                if cf_fill_xml not in fill_additions:
                    styles_xml, fidx = _ensure_fill(styles_xml, fill_xml_for_cell)
                    fill_additions[cf_fill_xml] = fidx
                fill_idx = fill_additions[cf_fill_xml]

                # Get original cell xf id
                cell_el = re.search(
                    rf'<c r="{ref}"[^>]*>',
                    sheet_xml)
                if cell_el:
                    s_m = re.search(r's="(\d+)"', cell_el.group(0))
                    orig_xf = int(s_m.group(1)) if s_m else 0

                    xf_key = (orig_xf, fill_idx)
                    if xf_key not in new_xf_fill_map:
                        styles_xml, new_xf = _ensure_cellxf(
                            styles_xml, orig_xf, fill_idx, original_fill_count)
                        new_xf_fill_map[xf_key] = new_xf
                        styles_modified = True
                    cell_new_style[ref] = new_xf_fill_map[xf_key]

    if not cell_new_index:
        print("Nothing to change — no mapped lines found in the target week.")
        return

    # ---------- update sharedStrings ----------
    total = int(re.search(r'count="(\d+)"', sst_xml).group(1))
    sst_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{n_si + len(new_sis)}"',
                     sst_xml, count=1)
    sst_xml = re.sub(r'count="\d+"', f'count="{total + len(cell_new_index)}"',
                     sst_xml, count=1)
    sst_xml = sst_xml.replace("</sst>", "".join(new_sis) + "</sst>")

    # ---------- repoint cells (SI) ----------
    for ref, new_idx in cell_new_index.items():
        pat = re.compile(r'(<c r="%s"[^>]*t="s"[^>]*>.*?<v>)\d+(</v>)' % ref, re.S)
        sheet_xml, n = pat.subn(r"\g<1>%d\g<2>" % new_idx, sheet_xml, count=1)
        if n != 1:
            raise RuntimeError(f"cell {ref}: expected 1 SI replacement, got {n}")

    # ---------- repoint cells (s/style) for CF fill ----------
    for ref, new_xf in cell_new_style.items():
        # Replace s="..." in the <c> tag
        def _replace_s(m):
            full = m.group(0)
            if 's="' in full:
                return re.sub(r's="\d+"', f's="{new_xf}"', full, count=1)
            else:
                # inject s attribute
                return re.sub(r'(<c\s+r="[^"]+")', rf'\1 s="{new_xf}"', full, count=1)
        pat = re.compile(rf'<c r="{ref}"[^>]*>', re.S)
        sheet_xml = pat.sub(_replace_s, sheet_xml, count=1)

    # ---------- hide rows ----------
    sheet_xml = _hide_rows(sheet_xml, week_row)

    # ---------- write output ----------
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "xl/sharedStrings.xml":
                zout.writestr(item, sst_xml)
            elif item.filename == sheet_name:
                zout.writestr(item, sheet_xml)
            elif item.filename == "xl/styles.xml" and styles_modified:
                zout.writestr(item, styles_xml)
            else:
                zout.writestr(item, zin.read(item.filename))

    # Count hidden rows
    hidden_count = len(re.findall(r'hidden="1"', sheet_xml))
    print(f"OK: {len(cell_new_index)} cells updated, "
          f"{len(cell_new_style)} CF fills baked, "
          f"{hidden_count} rows hidden -> {dst}")


if __name__ == "__main__":
    main()
