#!/usr/bin/env python3
"""Apply a Chinese translation mapping to one week of the flight-test planning
workbook.  Edits the .xlsm via direct XML surgery (zip-level) so VBA macros,
images, and data validation survive byte-for-byte.  The ORIGINAL file is never
modified — output goes to a new path (the skill names it ``*_CH_EN.xlsm``).

Usage:
    python apply_translation.py <in.xlsm> <context.json> <mapping.json> <out.xlsm>

  * context.json  — written by extract_week.py: {"sheet","week","week_row",
    "cells": {"B267": <si_index>, ...}}.
  * mapping.json  — the ``{"to_translate": {english: chinese}}`` payload with
    every value filled in (a bare {english: chinese} dict is also accepted).

Colours (v4):
  Conditional-formatting backgrounds/fonts are NOT re-derived from the workbook
  at run time.  They are read from the static ``references/cf_colors.json``
  table (regenerate with ``dump_cf_colors.py`` if the workbook's CF changes).
  For each translated cell whose full text matches a table key, the stored
  ``fill`` (a ready ``<patternFill>``) is baked into a new cellXf and the stored
  ``font`` colour is inherited by the Chinese run.  This keeps every run
  deterministic (better prompt-cache hits) and cheap (no CF data in the
  extract JSON, no theme/indexed resolution per cell).

Font consistency: Chinese runs inherit the English run's colour, font name and
size (capped at 9 pt); bold is stripped.

Row hiding: after translation ALL rows except header rows 1-3 and the ENTIRE
10-row week block (week_row … week_row+9) are hidden.  For multi-week chains the
caller MUST do a final visibility pass (post_process.py) to unhide the rows of
ALL translated weeks — this step only keeps its own week visible.
"""

import json
import os
import re
import sys
import zipfile
from xml.sax.saxutils import escape

# ---------------------------------------------------------------------------
# shared constants / paths
# ---------------------------------------------------------------------------
M_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
# Inspection line: "<N> FH Inspection <suffix>" OR "<N> Month(s) Inspection <suffix>".
# group(1)=number, group(2)=unit (FH|Month|Months), group(3)=suffix (qualifier) if any.
INSP_RE = re.compile(r"^\s*(\d+)\s*(FH|Months?)\s*Inspection\b(.*)$", re.I)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CF_COLORS_PATH = os.path.join(_SCRIPT_DIR, "..", "references", "cf_colors.json")

# CF rule: ``cellIs equal`` whose formula is a double-quoted literal like
# ``"Flight Test"`` (used only by dump_cf_colors.py, kept here so that helper
# imports resolve from one place).
CF_FORMULA_RE = re.compile(r'^"([^"]*)"$')


# ---------------------------------------------------------------------------
# XML helpers
# ---------------------------------------------------------------------------
def _sheet_xml_name(z, sheet):
    wb = z.read("xl/workbook.xml").decode()
    rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb))[sheet]
    rels = z.read("xl/_rels/workbook.xml.rels").decode()
    tgt = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))[rid]
    return "xl/" + tgt.lstrip("/")


def _unescape(s):
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'").replace("&amp;", "&"))


def _col_number(col_ref):
    n = 0
    for ch in col_ref:
        n = n * 26 + ord(ch.upper()) - 64
    return n


# ---------------------------------------------------------------------------
# static colour table
# ---------------------------------------------------------------------------
def load_cf_colors(path=_CF_COLORS_PATH):
    """Load the static activity-text -> {fill, font} table (see cf_colors.json).

    Returns {} if the file is missing so translation still works (colours only
    are skipped).
    """
    try:
        with open(path, encoding="utf-8-sig") as fh:
            return json.load(fh).get("map", {})
    except FileNotFoundError:
        sys.stderr.write(f"WARNING: {path} not found — CF colours will not be baked.\n")
        return {}


def lookup_color(color_map, cell_text):
    """Return (fill_xml_or_None, font_argb_or_None) for a cell's full text."""
    entry = color_map.get(cell_text)
    if entry is None:
        entry = color_map.get(cell_text.strip())
    if entry is None:
        return None, None
    return entry.get("fill"), entry.get("font")


# ---------------------------------------------------------------------------
# shared-string parsing / building
# ---------------------------------------------------------------------------
def parse_si(si_xml):
    """Return list of (text, rPr_xml_or_None)."""
    runs = []
    if "<r>" not in si_xml and "<r " not in si_xml:
        m = re.search(r"<t[^>]*>(.*?)</t>", si_xml, re.S)
        if m:
            runs.append((_unescape(m.group(1)), None))
        return runs
    for rm in re.finditer(r"<r>(.*?)</r>", si_xml, re.S):
        body = rm.group(1)
        pr = re.search(r"(<rPr>.*?</rPr>)", body, re.S)
        tm = re.search(r"<t[^>]*>(.*?)</t>", body, re.S)
        runs.append((_unescape(tm.group(1)) if tm else "", pr.group(1) if pr else None))
    return runs


def runs_to_lines(runs):
    """Split runs at newlines -> list of lines, each list of (text, rPr)."""
    lines, cur = [], []
    for text, pr in runs:
        parts = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        for i, part in enumerate(parts):
            if i > 0:
                lines.append(cur)
                cur = []
            if part:
                cur.append((part, pr))
    lines.append(cur)
    return lines


def _strip_bold(rpr_xml):
    if not rpr_xml:
        return None
    return re.sub(r"<b\s*/>", "", rpr_xml)


def _cap_sz(rpr_xml, max_sz=9):
    """Cap <sz val="N"/> so N ≤ max_sz.  Return (rpr_xml, changed)."""
    if not rpr_xml:
        return rpr_xml, False
    m = re.search(r'<sz\s+val="([\d.]+)"', rpr_xml)
    if not m:
        return rpr_xml, False
    cur = float(m.group(1))
    if cur <= max_sz:
        return rpr_xml, False
    new = re.sub(r'(<sz\s+val=")[\d.]+(")', rf'\g<1>{max_sz:g}\g<2>', rpr_xml, count=1)
    return new, True


def _cn_rpr(line_runs, default_rpr, cf_font_color=None, cell_colour=None, max_size=9):
    """Build rPr string for the Chinese run from the line's first styled run.

    Priority: explicit run rPr → CF font colour → resolved cell default colour.
    Bold is always stripped and font size capped at *max_size* pts.

    Font name, size, and family are inherited from *default_rpr* when the
    source run rPr doesn't provide them, so Chinese text always has explicit
    formatting that matches the English cell font.
    """
    src_rpr = None
    for _, pr in line_runs:
        if pr:
            src_rpr = pr
            break
    if not src_rpr:
        src_rpr = default_rpr

    rpr = _strip_bold(src_rpr)

    # Inherit font name from default_rpr if source rPr doesn't specify one
    if default_rpr and '<rFont' not in (rpr or ''):
        m = re.search(r'<rFont[^>]*/>', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0), 1)
            else:
                rpr = f'<rPr>{m.group(0)}</rPr>'

    # Inherit font family from default_rpr if missing
    if default_rpr and '<family' not in (rpr or ''):
        m = re.search(r'<family\s+val="[^"]*"', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0) + '/>', 1)
            else:
                rpr = f'<rPr>{m.group(0)}/></rPr>'

    # Inherit font size from default_rpr if missing (before CF colour
    # injection so the size element comes first)
    if default_rpr and '<sz ' not in (rpr or ''):
        m = re.search(r'<sz\s+val="[\d.]+"', default_rpr)
        if m:
            if rpr:
                rpr = rpr.replace('<rPr>', '<rPr>' + m.group(0) + '/>', 1)
            else:
                rpr = f'<rPr>{m.group(0)}/></rPr>'

    # Colour: explicit run colour > CF colour > resolved cell default colour
    # > black.  An explicit colour is ALWAYS baked into the Chinese run (never
    # left to inheritance).  Once a plain-string cell is converted to rich text,
    # English runs that carry no <color> would render black, so the resolved
    # ORIGINAL cell-font colour is baked in instead — Chinese matches the
    # original English display colour exactly.
    if '<color' not in (rpr or ''):
        colour = cf_font_color or cell_colour or 'FF000000'
        if rpr:
            rpr = rpr.replace('<rPr>', f'<rPr><color rgb="{colour}"/>', 1)
        else:
            rpr = f'<rPr><color rgb="{colour}"/></rPr>'

    rpr, _ = _cap_sz(rpr, max_size)
    return rpr


def build_si(lines_out):
    """lines_out: list of lines, each list of (text, rPr).  Return <si> XML."""
    parts = ["<si>"]
    flat = []
    for i, line in enumerate(lines_out):
        if i > 0:
            flat.append(("\r\n", line[0][1] if line else None))
        flat.extend(line if line else [])
    # merge adjacent runs with identical rPr
    merged = []
    for text, pr in flat:
        if merged and merged[-1][1] == pr:
            merged[-1] = (merged[-1][0] + text, pr)
        else:
            merged.append((text, pr))
    for text, pr in merged:
        t = f'<t xml:space="preserve">{escape(text)}</t>'
        parts.append(f"<r>{pr or ''}{t}</r>")
    parts.append("</si>")
    return "".join(parts)


# ---------------------------------------------------------------------------
# DXF / theme / indexed resolution — used ONLY by dump_cf_colors.py to build the
# static table offline.  Not called during translation.
# ---------------------------------------------------------------------------
def _parse_dxfs(styles_xml):
    """Return list of dicts {fill_xml, font_color_rgb, font_theme,
    font_color_indexed, bold, font_xml}."""
    dxfs = re.findall(r"<dxf>(.*?)</dxf>", styles_xml, re.S)
    parsed = []
    for d in dxfs:
        entry = {"fill_xml": None, "font_color_rgb": None,
                 "font_theme": None, "font_color_indexed": None,
                 "bold": False, "font_xml": None}
        fill_m = re.search(r"<fill>(.*?)</fill>", d, re.S)
        if fill_m:
            entry["fill_xml"] = fill_m.group(1).strip() or None
        font_m = re.search(r"<font>(.*?)</font>", d, re.S)
        if font_m:
            font_body = font_m.group(1)
            entry["font_xml"] = font_body
            entry["bold"] = "<b/>" in font_body
            color_m = re.search(r'<color\s+(.*?)/?>', font_body)
            if color_m:
                cattrs = color_m.group(1)
                crgb = re.search(r'rgb="([^"]*)"', cattrs)
                ctheme = re.search(r'theme="([^"]*)"', cattrs)
                cidx = re.search(r'indexed="([^"]*)"', cattrs)
                if crgb:
                    entry["font_color_rgb"] = crgb.group(1)
                if ctheme:
                    entry["font_theme"] = ctheme.group(1)
                if cidx:
                    entry["font_color_indexed"] = cidx.group(1)
        parsed.append(entry)
    return parsed


def _parse_cf_rules(sheet_xml):
    """Return list of (sqref_parts, list_of_rule_dicts) for ``cellIs equal``
    rules with a literal ``"text"`` formula."""
    blocks = re.findall(
        r'<conditionalFormatting\s+sqref="([^"]*)"\s*>(.*?)</conditionalFormatting>',
        sheet_xml, re.S)
    rules = []
    for sqref, body in blocks:
        parts = sqref.split()
        rule_list = []
        for attrs_str, rule_body in re.findall(r"<cfRule(.*?)>(.*?)</cfRule>", body, re.S):
            a = dict(re.findall(r'(\w+)="([^"]*)"', attrs_str))
            if a.get("type") != "cellIs" or a.get("operator") != "equal":
                continue
            f_m = re.search(r"<formula>(.*?)</formula>", rule_body)
            if not f_m:
                continue
            fm = CF_FORMULA_RE.match(f_m.group(1).strip())
            if not fm:
                continue
            rule_list.append({
                "text": fm.group(1),
                "dxfId": int(a.get("dxfId", "0")),
                "priority": int(a.get("priority", "0")),
            })
        if rule_list:
            rules.append((parts, rule_list))
    return rules


def _resolve_theme_colors(z):
    """Read theme1.xml from the zip and return dict {theme_idx: "RRGGBB"}."""
    try:
        theme_xml = z.read("xl/theme/theme1.xml").decode()
    except (KeyError, zipfile.BadZipFile):
        return {}
    scheme = re.search(r"<a:clrScheme[^>]*>(.*?)</a:clrScheme>", theme_xml, re.S)
    if not scheme:
        return {}
    theme_map = {
        "dk1": 1, "lt1": 0, "dk2": 3, "lt2": 2,
        "accent1": 4, "accent2": 5, "accent3": 6,
        "accent4": 7, "accent5": 8, "accent6": 9,
    }
    colors = {}
    for name, idx in theme_map.items():
        # Attribute order varies (e.g. <a:sysClr val="windowText" lastClr="000000"/>),
        # so match the colour value wherever it sits in the clr element.
        block = re.search(rf"<a:{name}>(.*?)</a:{name}>", scheme.group(1), re.S)
        if not block:
            continue
        m = re.search(r'<a:srgbClr[^>]*\bval="([^"]*)"', block.group(1))
        if not m:
            m = re.search(r'<a:sysClr[^>]*\blastClr="([^"]*)"', block.group(1))
        if m:
            colors[idx] = m.group(1).upper()
    return colors


def _resolve_indexed_colors(styles_xml):
    """Parse <indexedColors> from styles.xml, return {idx: 'RRGGBB'}."""
    colors = {}
    palette = re.search(r"<indexedColors>(.*?)</indexedColors>", styles_xml, re.S)
    if not palette:
        return colors
    entries = re.findall(r"<rgbColor[^>]*rgb=\"([^\"]+)\"", palette.group(1))
    for i, rgb in enumerate(entries):
        colors[i] = rgb.upper()
    return colors


def _build_fill_xml(rgb_color):
    return (
        '<patternFill patternType="solid">'
        f'<fgColor rgb="{rgb_color}"/>'
        '<bgColor indexed="64"/>'
        '</patternFill>'
    )


def _build_fill_direct(dxf_fill_xml):
    """DXF fill XML -> (ready <patternFill> XML, rgb_or_None).  Used by the
    generator; handles rgb, theme+tint and indexed bgColor sources."""
    bg_m = re.search(r'<bgColor\s*(.*?)/?>', dxf_fill_xml)
    if not bg_m:
        return None, None
    attrs = bg_m.group(1)
    rgb_m = re.search(r'rgb="([^"]*)"', attrs)
    if rgb_m:
        rgb = rgb_m.group(1)
        return _build_fill_xml(rgb), rgb
    theme_m = re.search(r'theme="([^"]*)"', attrs)
    if theme_m:
        tint_m = re.search(r'tint="([^"]*)"', attrs)
        tint = f' tint="{tint_m.group(1)}"' if tint_m else ""
        return (f'<patternFill patternType="solid">'
                f'<fgColor theme="{theme_m.group(1)}"{tint}/>'
                f'<bgColor indexed="64"/>'
                f'</patternFill>', None)
    indexed_m = re.search(r'indexed="([^"]*)"', attrs)
    if indexed_m:
        return (f'<patternFill patternType="solid">'
                f'<fgColor indexed="{indexed_m.group(1)}"/>'
                f'<bgColor indexed="64"/>'
                f'</patternFill>', None)
    return None, None


# ---------------------------------------------------------------------------
# fill / cellXf management
# ---------------------------------------------------------------------------
def _ensure_fill(styles_xml, fill_xml):
    """Append a fill to styles.xml if not already present; return (new_xml, fill_idx)."""
    fills_section = re.search(r"(<fills\s+count=\")(\d+)(\">)", styles_xml)
    if not fills_section:
        raise RuntimeError("Cannot locate <fills> in styles.xml")
    count = int(fills_section.group(2))
    existing = re.findall(r"<fill>(.*?)</fill>", styles_xml, re.S)
    for i, f in enumerate(existing):
        if f.strip() == fill_xml.strip():
            return styles_xml, i
    new_fill = f"<fill>{fill_xml}</fill>"
    styles_xml = styles_xml.replace(fills_section.group(0), f'<fills count="{count + 1}">', 1)
    idx = styles_xml.rfind("</fills>")
    styles_xml = styles_xml[:idx] + new_fill + styles_xml[idx:]
    return styles_xml, count


def _ensure_cellxf(styles_xml, original_xf_id, fill_idx):
    """Create a new cellXf that clones *original_xf_id* but with *fill_idx*.

    Returns (styles_xml, new_xf_id).
    """
    xf_section = re.search(r"(<cellXfs\s+count=\")(\d+)(\">)", styles_xml)
    xf_count = int(xf_section.group(2))
    xfs_block = re.search(r"<cellXfs[^>]*>(.*?)</cellXfs>", styles_xml, re.S)

    parts = re.split(r"(?=<xf\b)", xfs_block.group(1))
    xf_entries_raw = [p.strip() for p in parts if p.strip().startswith("<xf")]

    if original_xf_id >= len(xf_entries_raw):
        return styles_xml, original_xf_id

    orig = xf_entries_raw[original_xf_id]
    if 'fillId=' in orig:
        new_xf = re.sub(r'(fillId=\")\d+(\")', rf'\g<1>{fill_idx}\g<2>', orig, count=1)
    else:
        new_xf = orig.replace('<xf ', f'<xf fillId="{fill_idx}" ', 1)

    styles_xml = styles_xml.replace(xf_section.group(0), f'<cellXfs count="{xf_count + 1}">', 1)
    idx = styles_xml.rfind("</cellXfs>")
    styles_xml = styles_xml[:idx] + new_xf + styles_xml[idx:]
    return styles_xml, xf_count  # new xf id = old count (0-based)


# ---------------------------------------------------------------------------
# cell-level font / colour resolution
# ---------------------------------------------------------------------------
def _cell_default_colour_argb(style_id, cellxfs, fonts, indexed_map, theme_map):
    """Resolve a cell's default font colour to ``FFRRGGBB`` from styles.xml.

    Returns ``None`` when no colour is applied (the cell renders black).  This
    is the colour the ORIGINAL English text displays for lines whose runs carry
    no explicit colour — baking it into both the English and Chinese runs
    reproduces the original exactly after the plain-string → rich-text
    conversion.
    """
    if style_id is None or style_id >= len(cellxfs):
        return None
    xf = cellxfs[style_id]
    if re.search(r'applyFont="0"', xf):
        return None
    fm = re.search(r'fontId="(\d+)"', xf)
    fid = int(fm.group(1)) if fm else 0
    fb = fonts[fid] if fid < len(fonts) else fonts[0]
    cm = re.search(r'<color\s+(.*?)/>', fb)
    if not cm:
        return None
    a = cm.group(1)
    r = re.search(r'rgb="([^"]*)"', a)
    if r:
        v = r.group(1)
        return v if len(v) == 8 else 'FF' + v.upper()
    i = re.search(r'indexed="([^"]*)"', a)
    if i:
        v = indexed_map.get(int(i.group(1)))
        if v:
            v = v.upper()
            return v if len(v) == 8 else 'FF' + v
        return 'FF000000'
    t = re.search(r'theme="([^"]*)"', a)
    if t:
        v = theme_map.get(int(t.group(1)))
        if v:
            return 'FF' + v.upper()
    return None


def cell_default_rpr(cell_font, cf_font_color=None, cell_colour=None, max_size=9):
    """Build an rPr from an openpyxl Font, optionally overriding colour.

    ``cell_colour`` is the cell's true default font colour resolved from
    styles.xml (not openpyxl).  It is used when a Chinese line's English source
    run has no explicit colour, so the Chinese inherits exactly the colour the
    original English text displays.  A CF font colour (from cf_colors.json)
    still wins when set, so activity cells keep their conditional-format colour.

    Font size is capped at *max_size* (9 pt).
    """
    bits = []
    sz = cell_font.size if cell_font.size else 9
    sz = min(sz, max_size)
    bits.append(f'<sz val="{sz:g}"/>')
    colour = cf_font_color or cell_colour
    if colour:
        bits.append(f'<color rgb="{colour}"/>')
    if cell_font.name:
        bits.append(f'<rFont val="{cell_font.name}"/>')
        bits.append('<family val="2"/>')
    return "<rPr>" + "".join(bits) + "</rPr>" if bits else None


# ---------------------------------------------------------------------------
# line translation logic
# ---------------------------------------------------------------------------
def _line_colour(line_runs, cf_font_color=None, cell_colour=None):
    """Return a ready ``<color .../>`` element: the line's first explicit run
    colour if present, else the CF font colour, else the resolved cell default
    colour, else black.  This is the colour the ORIGINAL English line displays,
    baked explicitly into English + Chinese runs so both always match it."""
    for _, pr in line_runs:
        if pr:
            cm = re.search(r'<color\s+[^>]*?/>', pr)
            if cm:
                return cm.group(0)
    if cf_font_color:
        return f'<color rgb="{cf_font_color}"/>'
    if cell_colour:
        return f'<color rgb="{cell_colour}"/>'
    return '<color rgb="FF000000"/>'


def _bake_colour(line, colour_xml):
    """Return *line* with *colour_xml* added to every run lacking a colour.

    Runs that already carry an explicit colour are left untouched, so each
    English line keeps its own original colour."""
    out = []
    for text, pr in line:
        if pr and '<color' in pr:
            out.append((text, pr))
        elif pr:
            out.append((text, pr.replace('<rPr>', '<rPr>' + colour_xml, 1)))
        else:
            out.append((text, f'<rPr>{colour_xml}</rPr>'))
    return out


def translate_lines(lines, mapping, default_rpr, cf_font_color=None, cell_colour=None, cn_max_sz=9):
    """Insert Chinese translations; return (new_lines, changed?)."""
    out, changed, i = [], False, 0
    n = len(lines)
    while i < n:
        text_i = "".join(t for t, _ in lines[i])
        m = INSP_RE.match(text_i)
        if m:
            # Merge only consecutive lines with the SAME unit type AND SAME
            # suffix (precision).  Lines whose qualifiers differ are NOT merged.
            unit0 = m.group(2)
            unit0_norm = unit0.lower().rstrip("s")
            nums = [m.group(1)]
            suffix = m.group(3).strip()
            j = i + 1
            while j < n:
                tj = "".join(t for t, _ in lines[j])
                mm = INSP_RE.match(tj)
                if not mm:
                    break
                if mm.group(2).lower().rstrip("s") != unit0_norm:
                    break
                if mm.group(3).strip() != suffix:
                    break
                nums.append(mm.group(1))
                j += 1
            key = "/".join(nums) + unit0 + " Inspection" + (f" {suffix}" if suffix else "")
            colour_xml = _line_colour(lines[i], cf_font_color, cell_colour)
            for k in range(i, j):
                out.append(_bake_colour(lines[k], colour_xml))
            cn = mapping.get(key, "").strip()
            if cn:
                cn_pr = _cn_rpr(lines[j - 1], default_rpr, cf_font_color, cell_colour, cn_max_sz)
                out.append([(cn, cn_pr)])
                changed = True
            i = j
        else:
            colour_xml = _line_colour(lines[i], cf_font_color, cell_colour)
            out.append(_bake_colour(lines[i], colour_xml))
            cn = mapping.get(text_i.strip(), "").strip()
            if cn:
                cn_pr = _cn_rpr(lines[i], default_rpr, cf_font_color, cell_colour, cn_max_sz)
                out.append([(cn, cn_pr)])
                changed = True
            i += 1
    return out, changed


# ---------------------------------------------------------------------------
# row hiding
# ---------------------------------------------------------------------------
def _hide_rows(sheet_xml, week_row):
    """Hide ALL rows except header rows 1-3 and the ENTIRE 10-row block of the
    target week (week_row through week_row+9)."""
    visible_rows = {1, 2, 3} | set(range(week_row, week_row + 10))

    def _maybe_hide(m):
        r = int(m.group(1))
        if r not in visible_rows:
            if 'hidden="1"' not in m.group(0):
                return re.sub(r'(<row\b[^>]*)>', r'\1 hidden="1">', m.group(0), count=1)
        else:
            if 'hidden="1"' in m.group(0):
                return m.group(0).replace(' hidden="1"', '')
        return m.group(0)

    sheet_xml = re.sub(r'<row\b[^>]*r="(\d+)"[^>]*>', _maybe_hide, sheet_xml)
    return sheet_xml


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def _load_mapping(map_path):
    data = json.load(open(map_path, encoding="utf-8-sig"))
    if isinstance(data, dict) and "to_translate" in data:
        return data["to_translate"]
    return data  # already a flat {english: chinese} dict


def main():
    if len(sys.argv) != 5:
        sys.exit("Usage: apply_translation.py <in.xlsm> <context.json> "
                 "<mapping.json> <out.xlsm>")
    src, ctx_path, map_path, dst = sys.argv[1:5]

    ctx = json.load(open(ctx_path, encoding="utf-8-sig"))
    sheet = ctx["sheet"]
    week_row = ctx["week_row"]
    cells = ctx["cells"]                     # {ref: si_index}
    mapping = _load_mapping(map_path)
    color_map = load_cf_colors()

    # ---------- openpyxl pass (READ-ONLY, never save) ----------
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(src, read_only=False, keep_vba=True)
    ws = wb[sheet]

    # ---------- read zip internals ----------
    zin = zipfile.ZipFile(src)
    sheet_name = _sheet_xml_name(zin, sheet)
    sst_xml = zin.read("xl/sharedStrings.xml").decode()
    sheet_xml = zin.read(sheet_name).decode()
    styles_xml = zin.read("xl/styles.xml").decode()

    si_bodies = re.findall(r"<si>.*?</si>", sst_xml, re.S)
    n_si = len(si_bodies)

    # ---------- resolve cell default font colours from styles.xml ----------
    indexed_map = _resolve_indexed_colors(styles_xml)
    theme_map = _resolve_theme_colors(zin)
    _fonts = re.findall(r"<font>(.*?)</font>", styles_xml, re.S)
    _cellxfs_block = re.search(r"<cellXfs[^>]*>(.*?)</cellXfs>", styles_xml, re.S)
    _cellxfs = [x for x in re.split(r"(?=<xf\b)", _cellxfs_block.group(1))
                if x.strip().startswith("<xf")]

    new_sis, si_cache = [], {}
    cell_new_index = {}           # ref -> new SI index
    cell_new_style = {}           # ref -> new cellXf id (when CF fill baked)
    new_xf_fill_map = {}          # (original_xfId, fill_idx) -> new_xfId
    fill_additions = {}           # fill_xml -> fill_idx
    styles_modified = False

    # ---------- process each cell ----------
    for ref, si_info in cells.items():
        si_idx = si_info if isinstance(si_info, int) else si_info["si"]
        runs = parse_si(si_bodies[si_idx])
        lines = runs_to_lines(runs)
        cell_text = "".join(r[0] for r in runs)   # full text, for colour lookup

        # ---------- colour from static table ----------
        cf_fill_xml, cf_font_color = lookup_color(color_map, cell_text)

        # ---------- inject CF font colour into English runs ----------
        # After translation the cell text changes (bilingual), so the CF
        # "cellIs equal" rule no longer matches.  English runs that relied
        # on CF for their font colour would lose it.  Pre-inject the colour
        # so both English and Chinese runs keep the correct colour.
        if cf_font_color:
            for line in lines:
                for i, (text, rpr) in enumerate(line):
                    if '<color' not in (rpr or ''):
                        if rpr:
                            line[i] = (text, rpr.replace(
                                '<rPr>', f'<rPr><color rgb="{cf_font_color}"/>', 1))
                        else:
                            line[i] = (text,
                                       f'<rPr><color rgb="{cf_font_color}"/></rPr>')

        cell_colour = _cell_default_colour_argb(ws[ref].style_id, _cellxfs, _fonts,
                                                indexed_map, theme_map)
        default = cell_default_rpr(ws[ref].font, cf_font_color, cell_colour, max_size=9)
        new_lines, changed = translate_lines(lines, mapping, default, cf_font_color,
                                             cell_colour, cn_max_sz=9)
        if not changed:
            continue

        key = (si_idx, default)
        if key not in si_cache:
            si_cache[key] = n_si + len(new_sis)
            new_sis.append(build_si(new_lines))
        cell_new_index[ref] = si_cache[key]

        # ---------- bake CF fill into a cellXf ----------
        if cf_fill_xml:
            if cf_fill_xml not in fill_additions:
                styles_xml, fidx = _ensure_fill(styles_xml, cf_fill_xml)
                fill_additions[cf_fill_xml] = fidx
            fill_idx = fill_additions[cf_fill_xml]

            cell_el = re.search(rf'<c r="{ref}"[^>]*>', sheet_xml)
            if cell_el:
                s_m = re.search(r's="(\d+)"', cell_el.group(0))
                orig_xf = int(s_m.group(1)) if s_m else 0
                xf_key = (orig_xf, fill_idx)
                if xf_key not in new_xf_fill_map:
                    styles_xml, new_xf = _ensure_cellxf(styles_xml, orig_xf, fill_idx)
                    new_xf_fill_map[xf_key] = new_xf
                    styles_modified = True
                cell_new_style[ref] = new_xf_fill_map[xf_key]

    if not cell_new_index:
        print("Nothing to change — no mapped lines found in the target week.")
        return

    # ---------- update sharedStrings ----------
    total = int(re.search(r'count="(\d+)"', sst_xml).group(1))
    sst_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{n_si + len(new_sis)}"',
                     sst_xml, count=1)
    sst_xml = re.sub(r'count="\d+"', f'count="{total + len(cell_new_index)}"',
                     sst_xml, count=1)
    sst_xml = sst_xml.replace("</sst>", "".join(new_sis) + "</sst>")

    # ---------- repoint cells (SI) ----------
    for ref, new_idx in cell_new_index.items():
        pat = re.compile(r'(<c r="%s"[^>]*t="s"[^>]*>.*?<v>)\d+(</v>)' % ref, re.S)
        sheet_xml, n = pat.subn(r"\g<1>%d\g<2>" % new_idx, sheet_xml, count=1)
        if n != 1:
            raise RuntimeError(f"cell {ref}: expected 1 SI replacement, got {n}")

    # ---------- repoint cells (s/style) for CF fill ----------
    for ref, new_xf in cell_new_style.items():
        def _replace_s(m):
            full = m.group(0)
            if 's="' in full:
                return re.sub(r's="\d+"', f's="{new_xf}"', full, count=1)
            return re.sub(r'(<c\s+r="[^"]+")', rf'\1 s="{new_xf}"', full, count=1)
        pat = re.compile(rf'<c r="{ref}"[^>]*>', re.S)
        sheet_xml = pat.sub(_replace_s, sheet_xml, count=1)

    # ---------- hide rows ----------
    sheet_xml = _hide_rows(sheet_xml, week_row)

    # ---------- write output (original file untouched) ----------
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "xl/sharedStrings.xml":
                zout.writestr(item, sst_xml)
            elif item.filename == sheet_name:
                zout.writestr(item, sheet_xml)
            elif item.filename == "xl/styles.xml" and styles_modified:
                zout.writestr(item, styles_xml)
            else:
                zout.writestr(item, zin.read(item.filename))

    hidden_count = len(re.findall(r'hidden="1"', sheet_xml))
    print(f"OK: {len(cell_new_index)} cells updated, "
          f"{len(cell_new_style)} CF fills baked, "
          f"{hidden_count} rows hidden -> {dst}")


if __name__ == "__main__":
    main()
